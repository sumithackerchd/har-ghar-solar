import io
import logging
import os
import secrets
import sqlite3
import smtplib
from datetime import datetime, timedelta
from email.message import EmailMessage
from functools import wraps
from flask import send_from_directory


from dotenv import load_dotenv
from flask import (
    Flask, render_template, request, redirect,
    session, send_file, flash, jsonify, abort, url_for
)
from flask_sqlalchemy import SQLAlchemy
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from werkzeug.security import generate_password_hash, check_password_hash
import pandas as pd

load_dotenv()

# ==========================
# LOGGING
# ==========================
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S"
)
logger = logging.getLogger("hargharsolar")

# ==========================
# UTTAR PRADESH 75 DISTRICTS
# ==========================
UP_DISTRICTS = [
    "Agra", "Aligarh", "Ambedkar Nagar", "Amethi", "Amroha",
    "Auraiya", "Ayodhya", "Azamgarh", "Baghpat", "Bahraich",
    "Ballia", "Balrampur", "Banda", "Barabanki", "Bareilly",
    "Basti", "Bhadohi", "Bijnor", "Budaun", "Bulandshahr",
    "Chandauli", "Chitrakoot", "Deoria", "Etah", "Etawah",
    "Farrukhabad", "Fatehpur", "Firozabad", "Gautam Buddha Nagar",
    "Ghaziabad", "Ghazipur", "Gonda", "Gorakhpur", "Hamirpur",
    "Hapur", "Hardoi", "Hathras", "Jalaun", "Jaunpur",
    "Jhansi", "Kannauj", "Kanpur Dehat", "Kanpur Nagar",
    "Kasganj", "Kaushambi", "Kushinagar", "Lakhimpur Kheri",
    "Lalitpur", "Lucknow", "Maharajganj", "Mahoba", "Mainpuri",
    "Mathura", "Mau", "Meerut", "Mirzapur", "Moradabad",
    "Muzaffarnagar", "Pilibhit", "Pratapgarh", "Prayagraj",
    "Rae Bareli", "Rampur", "Saharanpur", "Sambhal",
    "Sant Kabir Nagar", "Shahjahanpur", "Shamli", "Shravasti",
    "Siddharthnagar", "Sitapur", "Sonbhadra", "Sultanpur",
    "Unnao", "Varanasi"
]

LEAD_STATUSES = [
    "New", "Assigned", "Contacted",
    "Site Visit", "Quotation Sent", "Installation",
    "Completed", "Cancelled"
]

# ==========================
# APP CONFIG
# ==========================
app = Flask(__name__)
app.secret_key = os.getenv("SECRET_KEY", secrets.token_hex(32))
app.config["SQLALCHEMY_DATABASE_URI"] = os.getenv("DATABASE_URL", "sqlite:///solar.db")
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False
app.config["SQLALCHEMY_ENGINE_OPTIONS"] = {
    "pool_pre_ping": True,
    "pool_recycle": 300,
}
app.config["SESSION_COOKIE_HTTPONLY"] = True
app.config["SESSION_COOKIE_SAMESITE"] = "Lax"

db = SQLAlchemy(app)

# Jinja2 custom filter: parse JSON in templates
import json as _json_mod
app.jinja_env.filters['fromjson'] = lambda s: _json_mod.loads(s or '[]')

# ==========================
# RATE LIMITING (in-memory)
# ==========================
_rate_store: dict = {}

def rate_limit(max_calls: int, period: int):
    """Simple per-IP rate limiter decorator."""
    def decorator(f):
        @wraps(f)
        def wrapped(*args, **kwargs):
            ip  = request.remote_addr or "0.0.0.0"
            key = f"{f.__name__}:{ip}"
            now = datetime.now().timestamp()
            hits = [t for t in _rate_store.get(key, []) if now - t < period]
            if len(hits) >= max_calls:
                return render_template("errors/429.html"), 429
            hits.append(now)
            _rate_store[key] = hits
            return f(*args, **kwargs)
        return wrapped
    return decorator

# ==========================
# AUTH HELPERS
# ==========================
def login_required(f):
    @wraps(f)
    def wrapped(*args, **kwargs):
        if "user_id" not in session:
            return redirect(url_for("admin_login"))
        return f(*args, **kwargs)
    return wrapped

def admin_required(f):
    @wraps(f)
    def wrapped(*args, **kwargs):
        if "user_id" not in session:
            return redirect(url_for("admin_login"))
        if session.get("role") != "admin":
            abort(403)
        return f(*args, **kwargs)
    return wrapped

def super_admin_required(f):
    @wraps(f)
    def wrapped(*args, **kwargs):
        if "user_id" not in session:
            return redirect(url_for("admin_login"))
        if session.get("role") != "admin":
            abort(403)
        user = User.query.get(session["user_id"])
        if not user or not user.is_super:
            abort(403)
        return f(*args, **kwargs)
    return wrapped

def vendor_required(f):
    @wraps(f)
    def wrapped(*args, **kwargs):
        if "vendor_id" not in session:
            return redirect(url_for("vendor_login"))
        return f(*args, **kwargs)
    return wrapped

# ==========================
# EMAIL FUNCTION
# ==========================
def send_lead_email(lead):
    try:
        msg = EmailMessage()
        msg["Subject"] = "☀ New Solar Lead Received"
        msg["From"] = os.getenv("MAIL_USERNAME")
        msg["To"] = os.getenv("ADMIN_EMAIL")
        msg.set_content(f"""
New Solar Enquiry

Customer Name : {lead.name}
Phone Number  : {lead.phone}
City          : {lead.city}
District      : {lead.district}
Bill Range    : {lead.bill}
Date          : {lead.created_at}
        """)
        with smtplib.SMTP("smtp.gmail.com", 587) as server:
            server.starttls()
            server.login(os.getenv("MAIL_USERNAME"), os.getenv("MAIL_PASSWORD"))
            server.send_message(msg)
        print("EMAIL SENT SUCCESSFULLY")
    except Exception as e:
        print("EMAIL ERROR:", e)

# ==========================
# MODELS
# ==========================
class Vendor(db.Model):
    id            = db.Column(db.Integer, primary_key=True)
    company_name  = db.Column(db.String(150), nullable=False)
    owner_name    = db.Column(db.String(100), nullable=False)
    mobile        = db.Column(db.String(20), nullable=False)
    email         = db.Column(db.String(150), default="")
    username      = db.Column(db.String(50), unique=True, nullable=False)
    password      = db.Column(db.String(300), nullable=False)
    district      = db.Column(db.String(100), nullable=False)
    is_active     = db.Column(db.Boolean, default=True)
    is_deleted    = db.Column(db.Boolean, default=False)   # soft-delete flag
    created_at    = db.Column(db.String(50), default="")
    vendor_code   = db.Column(db.String(50), default="")
    address       = db.Column(db.Text, default="")
    approved_date = db.Column(db.String(50), default="")
    admin_remarks = db.Column(db.Text, default="")         # admin-only notes


class Lead(db.Model):
    id         = db.Column(db.Integer, primary_key=True)
    name       = db.Column(db.String(100))
    phone      = db.Column(db.String(20))
    city       = db.Column(db.String(100))
    bill       = db.Column(db.String(100))
    created_at = db.Column(db.String(50))
    status     = db.Column(db.String(50), default="New")
    note       = db.Column(db.Text, default="")
    follow_date = db.Column(db.String(50), default="")
    updated_by  = db.Column(db.String(100), default="")
    district    = db.Column(db.String(100), default="")
    vendor_id   = db.Column(db.Integer, db.ForeignKey("vendor.id"), nullable=True)
    vendor      = db.relationship("Vendor", backref="leads", foreign_keys=[vendor_id])


class LeadTimeline(db.Model):
    id         = db.Column(db.Integer, primary_key=True)
    lead_id    = db.Column(db.Integer, db.ForeignKey("lead.id"), nullable=False)
    event      = db.Column(db.String(200), nullable=False)
    note       = db.Column(db.Text, default="")
    created_by = db.Column(db.String(100), default="")
    created_at = db.Column(db.String(50), default="")
    lead       = db.relationship("Lead", backref="timeline")


class User(db.Model):
    id         = db.Column(db.Integer, primary_key=True)
    name       = db.Column(db.String(100))
    username   = db.Column(db.String(50), unique=True)
    password   = db.Column(db.String(300))
    role       = db.Column(db.String(50), default="employee")
    email      = db.Column(db.String(150), default="")
    is_super   = db.Column(db.Boolean, default=False)
    created_at = db.Column(db.String(50))


class VisitorCount(db.Model):
    id    = db.Column(db.Integer, primary_key=True)
    count = db.Column(db.Integer, default=0)



class Blog(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    title = db.Column(db.String(255), nullable=False)
    slug = db.Column(db.String(255), unique=True, nullable=False)
    content = db.Column(db.Text, nullable=False)
    meta_title = db.Column(db.String(255))
    meta_desc = db.Column(db.String(255))
    category = db.Column(db.String(100))
    tags = db.Column(db.String(255))
    author = db.Column(db.String(100), default="Admin")
    reading_time = db.Column(db.Integer, default=5)
    published_at = db.Column(db.String(50))
    updated_at = db.Column(db.String(50))
    featured_image = db.Column(db.String(300))
    schema_markup = db.Column(db.Text)
    is_published = db.Column(db.Boolean, default=True)

# ==========================
# VENDOR EXTENDED MODELS
# ==========================
class VendorProfile(db.Model):
    __tablename__ = "vendor_profile"
    id                   = db.Column(db.Integer, primary_key=True)
    vendor_id            = db.Column(db.Integer, db.ForeignKey("vendor.id"), unique=True, nullable=False)
    gst_number           = db.Column(db.String(20), default="")
    pan_number           = db.Column(db.String(20), default="")
    state                = db.Column(db.String(100), default="Uttar Pradesh")
    pincode              = db.Column(db.String(10), default="")
    alternate_phone      = db.Column(db.String(20), default="")
    website              = db.Column(db.String(200), default="")
    years_experience     = db.Column(db.Integer, default=0)
    install_capacity_kw  = db.Column(db.Integer, default=0)
    res_projects         = db.Column(db.Integer, default=0)
    comm_projects        = db.Column(db.Integer, default=0)
    ind_projects         = db.Column(db.Integer, default=0)
    govt_projects        = db.Column(db.Integer, default=0)
    pm_surya_approved    = db.Column(db.Boolean, default=False)
    discom_empanelled    = db.Column(db.Boolean, default=False)
    brands_supported     = db.Column(db.Text, default="")   # comma-separated
    company_logo         = db.Column(db.String(300), default="")
    profile_photo        = db.Column(db.String(300), default="")
    updated_at           = db.Column(db.String(50), default="")
    vendor               = db.relationship("Vendor", backref=db.backref("profile", uselist=False))


SOLAR_BRANDS = [
    "Waaree", "Adani", "Tata", "Luminous", "Havells",
    "Loom Solar", "Vikram Solar", "Premier Energies", "RenewSys", "Other"
]

LEAD_TYPES = [
    "Residential", "Commercial", "Government", "Industrial",
    "Society", "School", "Hospital", "Factory"
]

KW_CAPACITIES = [1, 2, 3, 5, 10, 15, 20, 25, 50]


class VendorPricing(db.Model):
    __tablename__ = "vendor_pricing"
    id                  = db.Column(db.Integer, primary_key=True)
    vendor_id           = db.Column(db.Integer, db.ForeignKey("vendor.id"), nullable=False)
    project_type        = db.Column(db.String(50), nullable=False)   # Residential/Commercial/Government
    capacity_kw         = db.Column(db.Integer, nullable=False)
    system_price        = db.Column(db.Float, default=0)             # type-specific base price
    brand               = db.Column(db.String(100), default="")
    warranty_years      = db.Column(db.Integer, default=0)
    vendor_price        = db.Column(db.Float, default=0)             # vendor selling price
    hgs_commission      = db.Column(db.Float, default=0)             # HGS commission
    final_price         = db.Column(db.Float, default=0)             # vendor_price + hgs_commission
    # Component-wise pricing
    panel_price         = db.Column(db.Float, default=0)   # per-unit panel price
    panel_wattage       = db.Column(db.Integer, default=0) # panel wattage in W
    panel_quantity      = db.Column(db.Integer, default=0) # number of panels
    inverter_price      = db.Column(db.Float, default=0)
    inverter_brand      = db.Column(db.String(100), default="")
    inverter_capacity   = db.Column(db.String(50), default="")  # e.g. "5 KW"
    structure_price     = db.Column(db.Float, default=0)
    structure_type      = db.Column(db.String(100), default="")
    install_charge      = db.Column(db.Float, default=0)
    elec_material       = db.Column(db.Float, default=0)
    net_meter_charge    = db.Column(db.Float, default=0)
    transportation      = db.Column(db.Float, default=0)
    documentation       = db.Column(db.Float, default=0)
    miscellaneous       = db.Column(db.Float, default=0)
    gst_percent         = db.Column(db.Float, default=0)   # GST %
    updated_at          = db.Column(db.String(50), default="")
    vendor              = db.relationship("Vendor", backref="pricing")
    __table_args__      = (db.UniqueConstraint("vendor_id", "project_type", "capacity_kw"),)


class VendorCommission(db.Model):
    __tablename__ = "vendor_commission"
    id               = db.Column(db.Integer, primary_key=True)
    vendor_id        = db.Column(db.Integer, db.ForeignKey("vendor.id"), nullable=False)
    project_type     = db.Column(db.String(50), nullable=False)
    capacity_kw      = db.Column(db.Integer, nullable=False)
    vendor_price     = db.Column(db.Float, default=0)
    hgs_commission   = db.Column(db.Float, default=0)
    final_price      = db.Column(db.Float, default=0)
    updated_at       = db.Column(db.String(50), default="")
    vendor           = db.relationship("Vendor", backref="commissions")
    __table_args__   = (db.UniqueConstraint("vendor_id", "project_type", "capacity_kw"),)


class VendorCoverage(db.Model):
    __tablename__ = "vendor_coverage"
    id            = db.Column(db.Integer, primary_key=True)
    vendor_id     = db.Column(db.Integer, db.ForeignKey("vendor.id"), nullable=False)
    districts     = db.Column(db.Text, default="")   # JSON list
    lead_types    = db.Column(db.Text, default="")   # JSON list
    all_districts = db.Column(db.Boolean, default=False)
    updated_at    = db.Column(db.String(50), default="")
    vendor        = db.relationship("Vendor", backref=db.backref("coverage", uselist=False))


class VendorQuotation(db.Model):
    __tablename__ = "vendor_quotation"
    id             = db.Column(db.Integer, primary_key=True)
    vendor_id      = db.Column(db.Integer, db.ForeignKey("vendor.id"), nullable=False)
    lead_id        = db.Column(db.Integer, db.ForeignKey("lead.id"), nullable=True)
    customer_name  = db.Column(db.String(150), default="")
    customer_phone = db.Column(db.String(20), default="")
    capacity_kw    = db.Column(db.Integer, default=0)
    project_type   = db.Column(db.String(50), default="Residential")
    brand          = db.Column(db.String(100), default="")
    panel_brand    = db.Column(db.String(100), default="")
    inverter_brand = db.Column(db.String(100), default="")
    structure_type = db.Column(db.String(100), default="")
    subsidy_amount = db.Column(db.Float, default=0)
    gross_price    = db.Column(db.Float, default=0)
    net_price      = db.Column(db.Float, default=0)
    commission     = db.Column(db.Float, default=0)
    remarks        = db.Column(db.Text, default="")
    status         = db.Column(db.String(30), default="Pending")
    admin_remarks  = db.Column(db.Text, default="")
    created_at     = db.Column(db.String(50), default="")
    updated_at     = db.Column(db.String(50), default="")
    # Item-wise breakdown stored as JSON
    components_json = db.Column(db.Text, default="")
    gst_amount      = db.Column(db.Float, default=0)
    sub_total       = db.Column(db.Float, default=0)
    gst_percent     = db.Column(db.Float, default=0)
    # ── Pricing Snapshot ─────────────────────────────────────────────────────
    # All fields copied from VendorPricing at quotation-creation time.
    # Quotation detail MUST read these — never re-query VendorPricing.
    snap_pricing_found   = db.Column(db.Boolean, default=False)
    snap_vendor_price    = db.Column(db.Float, default=0)
    snap_hgs_commission  = db.Column(db.Float, default=0)
    snap_final_price     = db.Column(db.Float, default=0)
    snap_brand           = db.Column(db.String(100), default="")
    snap_warranty_years  = db.Column(db.Integer, default=0)
    snap_panel_price     = db.Column(db.Float, default=0)
    snap_panel_wattage   = db.Column(db.Integer, default=0)
    snap_panel_quantity  = db.Column(db.Integer, default=0)
    snap_inverter_price  = db.Column(db.Float, default=0)
    snap_inverter_brand  = db.Column(db.String(100), default="")
    snap_inverter_cap    = db.Column(db.String(50), default="")
    snap_structure_price = db.Column(db.Float, default=0)
    snap_structure_type  = db.Column(db.String(100), default="")
    snap_install_charge  = db.Column(db.Float, default=0)
    snap_elec_material   = db.Column(db.Float, default=0)
    snap_net_meter       = db.Column(db.Float, default=0)
    snap_transportation  = db.Column(db.Float, default=0)
    snap_documentation   = db.Column(db.Float, default=0)
    snap_miscellaneous   = db.Column(db.Float, default=0)
    snap_gst_percent     = db.Column(db.Float, default=0)
    vendor               = db.relationship("Vendor", backref="quotations")
    lead                 = db.relationship("Lead", backref="quotations")


# ==========================
# VENDOR RATING MODEL
# ==========================
class VendorRating(db.Model):
    __tablename__ = "vendor_rating"
    id          = db.Column(db.Integer, primary_key=True)
    vendor_id   = db.Column(db.Integer, db.ForeignKey("vendor.id"), nullable=False)
    lead_id     = db.Column(db.Integer, db.ForeignKey("lead.id"), nullable=True)
    rating      = db.Column(db.Float, default=0)        # 1-5
    review      = db.Column(db.Text, default="")
    rated_by    = db.Column(db.String(100), default="")
    created_at  = db.Column(db.String(50), default="")
    vendor      = db.relationship("Vendor", backref="ratings")


# ==========================
# NOTIFICATION MODEL
# ==========================
class Notification(db.Model):
    __tablename__ = "notification"
    id          = db.Column(db.Integer, primary_key=True)
    target_type = db.Column(db.String(20), default="admin")  # admin / vendor
    target_id   = db.Column(db.Integer, nullable=True)       # vendor_id or user_id, None=all admins
    title       = db.Column(db.String(200), default="")
    message     = db.Column(db.Text, default="")
    is_read     = db.Column(db.Boolean, default=False)
    created_at  = db.Column(db.String(50), default="")

# ==========================
# DB INIT + MIGRATION
# ==========================
with app.app_context():
    db.create_all()

    # Safe migration for existing SQLite database
    db_path = os.path.join("instance", "solar.db")
    if os.path.exists(db_path):
        conn = sqlite3.connect(db_path)
        cur = conn.cursor()
        cur.execute("PRAGMA table_info(lead)")
        existing = [r[1] for r in cur.fetchall()]
        if "district" not in existing:
            cur.execute("ALTER TABLE lead ADD COLUMN district VARCHAR(100) DEFAULT ''")
        if "vendor_id" not in existing:
            cur.execute("ALTER TABLE lead ADD COLUMN vendor_id INTEGER")
        # User table migrations
        cur.execute("PRAGMA table_info(user)")
        user_cols = [r[1] for r in cur.fetchall()]
        if "is_super" not in user_cols:
            cur.execute("ALTER TABLE user ADD COLUMN is_super BOOLEAN DEFAULT 0")
        if "email" not in user_cols:
            cur.execute("ALTER TABLE user ADD COLUMN email VARCHAR(150) DEFAULT ''")
        # Vendor table migrations
        cur.execute("PRAGMA table_info(vendor)")
        vendor_cols = [r[1] for r in cur.fetchall()]
        if "vendor_code" not in vendor_cols:
            cur.execute("ALTER TABLE vendor ADD COLUMN vendor_code VARCHAR(50) DEFAULT ''")
        if "address" not in vendor_cols:
            cur.execute("ALTER TABLE vendor ADD COLUMN address TEXT DEFAULT ''")
        if "approved_date" not in vendor_cols:
            cur.execute("ALTER TABLE vendor ADD COLUMN approved_date VARCHAR(50) DEFAULT ''")
        if "is_deleted" not in vendor_cols:
            cur.execute("ALTER TABLE vendor ADD COLUMN is_deleted BOOLEAN DEFAULT 0")
        if "admin_remarks" not in vendor_cols:
            cur.execute("ALTER TABLE vendor ADD COLUMN admin_remarks TEXT DEFAULT ''")

        # VendorProfile migration
        cur.execute("PRAGMA table_info(vendor_profile)")
        vp_cols = [r[1] for r in cur.fetchall()]
        if not vp_cols:
            cur.execute("""CREATE TABLE IF NOT EXISTS vendor_profile (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                vendor_id INTEGER UNIQUE NOT NULL,
                gst_number VARCHAR(20) DEFAULT '',
                pan_number VARCHAR(20) DEFAULT '',
                state VARCHAR(100) DEFAULT 'Uttar Pradesh',
                pincode VARCHAR(10) DEFAULT '',
                alternate_phone VARCHAR(20) DEFAULT '',
                website VARCHAR(200) DEFAULT '',
                years_experience INTEGER DEFAULT 0,
                install_capacity_kw INTEGER DEFAULT 0,
                res_projects INTEGER DEFAULT 0,
                comm_projects INTEGER DEFAULT 0,
                ind_projects INTEGER DEFAULT 0,
                govt_projects INTEGER DEFAULT 0,
                pm_surya_approved BOOLEAN DEFAULT 0,
                discom_empanelled BOOLEAN DEFAULT 0,
                brands_supported TEXT DEFAULT '',
                company_logo VARCHAR(300) DEFAULT '',
                profile_photo VARCHAR(300) DEFAULT '',
                updated_at VARCHAR(50) DEFAULT '',
                FOREIGN KEY(vendor_id) REFERENCES vendor(id)
            )""")

        # VendorPricing migration
        cur.execute("""CREATE TABLE IF NOT EXISTS vendor_pricing (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            vendor_id INTEGER NOT NULL,
            project_type VARCHAR(50) NOT NULL,
            capacity_kw INTEGER NOT NULL,
            system_price FLOAT DEFAULT 0,
            install_charge FLOAT DEFAULT 0,
            gst_included BOOLEAN DEFAULT 1,
            brand VARCHAR(100) DEFAULT '',
            warranty_years INTEGER DEFAULT 0,
            structure_type VARCHAR(100) DEFAULT '',
            inverter_brand VARCHAR(100) DEFAULT '',
            panel_brand VARCHAR(100) DEFAULT '',
            battery VARCHAR(100) DEFAULT '',
            updated_at VARCHAR(50) DEFAULT '',
            FOREIGN KEY(vendor_id) REFERENCES vendor(id),
            UNIQUE(vendor_id, project_type, capacity_kw)
        )""")

        # VendorCommission migration
        cur.execute("""CREATE TABLE IF NOT EXISTS vendor_commission (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            vendor_id INTEGER NOT NULL,
            project_type VARCHAR(50) NOT NULL,
            capacity_kw INTEGER NOT NULL,
            vendor_price FLOAT DEFAULT 0,
            hgs_commission FLOAT DEFAULT 0,
            final_price FLOAT DEFAULT 0,
            updated_at VARCHAR(50) DEFAULT '',
            FOREIGN KEY(vendor_id) REFERENCES vendor(id),
            UNIQUE(vendor_id, project_type, capacity_kw)
        )""")

        # VendorCoverage migration
        cur.execute("""CREATE TABLE IF NOT EXISTS vendor_coverage (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            vendor_id INTEGER NOT NULL,
            districts TEXT DEFAULT '',
            lead_types TEXT DEFAULT '',
            all_districts BOOLEAN DEFAULT 0,
            updated_at VARCHAR(50) DEFAULT '',
            FOREIGN KEY(vendor_id) REFERENCES vendor(id)
        )""")

        # VendorQuotation migration
        cur.execute("""CREATE TABLE IF NOT EXISTS vendor_quotation (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            vendor_id INTEGER NOT NULL,
            lead_id INTEGER,
            customer_name VARCHAR(150) DEFAULT '',
            customer_phone VARCHAR(20) DEFAULT '',
            capacity_kw INTEGER DEFAULT 0,
            project_type VARCHAR(50) DEFAULT 'Residential',
            brand VARCHAR(100) DEFAULT '',
            panel_brand VARCHAR(100) DEFAULT '',
            inverter_brand VARCHAR(100) DEFAULT '',
            structure_type VARCHAR(100) DEFAULT '',
            subsidy_amount FLOAT DEFAULT 0,
            gross_price FLOAT DEFAULT 0,
            net_price FLOAT DEFAULT 0,
            commission FLOAT DEFAULT 0,
            remarks TEXT DEFAULT '',
            status VARCHAR(30) DEFAULT 'Pending',
            admin_remarks TEXT DEFAULT '',
            created_at VARCHAR(50) DEFAULT '',
            updated_at VARCHAR(50) DEFAULT '',
            FOREIGN KEY(vendor_id) REFERENCES vendor(id),
            FOREIGN KEY(lead_id) REFERENCES lead(id)
        )""")

        conn.commit()
        conn.close()

        # ── Re-open for new-column migrations ──────────────────────────
        conn2 = sqlite3.connect(db_path)
        cur2  = conn2.cursor()

        # VendorProfile — add missing columns
        cur2.execute("PRAGMA table_info(vendor_profile)")
        vp2 = [r[1] for r in cur2.fetchall()]
        new_vp_cols = [
            ("service_radius", "INTEGER DEFAULT 0"),
            ("bank_name",      "VARCHAR(100) DEFAULT ''"),
            ("account_number", "VARCHAR(50)  DEFAULT ''"),
            ("ifsc_code",      "VARCHAR(20)  DEFAULT ''"),
            ("account_holder", "VARCHAR(100) DEFAULT ''"),
            ("upi_id",         "VARCHAR(100) DEFAULT ''"),
        ]
        for col, defn in new_vp_cols:
            if col not in vp2:
                cur2.execute(f"ALTER TABLE vendor_profile ADD COLUMN {col} {defn}")

        # VendorPricing — add missing columns
        cur2.execute("PRAGMA table_info(vendor_pricing)")
        vpr2 = [r[1] for r in cur2.fetchall()]
        new_vpr_cols = [
            ("price_with_subsidy",    "FLOAT DEFAULT 0"),
            ("price_without_subsidy", "FLOAT DEFAULT 0"),
            ("govt_price",            "FLOAT DEFAULT 0"),
            ("panel_wattage",         "INTEGER DEFAULT 0"),
            ("install_days",          "INTEGER DEFAULT 0"),
            ("amc_available",         "BOOLEAN DEFAULT 0"),
            ("amc_cost_annual",       "FLOAT DEFAULT 0"),
            ("vendor_price",          "FLOAT DEFAULT 0"),
            ("hgs_commission",        "FLOAT DEFAULT 0"),
            ("final_price",           "FLOAT DEFAULT 0"),
            ("panel_price",           "FLOAT DEFAULT 0"),
            ("panel_quantity",        "INTEGER DEFAULT 0"),
            ("inverter_price",        "FLOAT DEFAULT 0"),
            ("inverter_brand",        "VARCHAR(100) DEFAULT ''"),
            ("inverter_capacity",     "VARCHAR(50) DEFAULT ''"),
            ("structure_price",       "FLOAT DEFAULT 0"),
            ("structure_type",        "VARCHAR(100) DEFAULT ''"),
            ("install_charge",        "FLOAT DEFAULT 0"),
            ("elec_material",         "FLOAT DEFAULT 0"),
            ("net_meter_charge",      "FLOAT DEFAULT 0"),
            ("transportation",        "FLOAT DEFAULT 0"),
            ("documentation",         "FLOAT DEFAULT 0"),
            ("miscellaneous",         "FLOAT DEFAULT 0"),
            ("gst_percent",           "FLOAT DEFAULT 0"),
        ]
        for col, defn in new_vpr_cols:
            if col not in vpr2:
                cur2.execute(f"ALTER TABLE vendor_pricing ADD COLUMN {col} {defn}")

        # VendorQuotation — add missing columns (including full pricing snapshot)
        cur2.execute("PRAGMA table_info(vendor_quotation)")
        vq2 = [r[1] for r in cur2.fetchall()]
        new_vq_cols = [
            ("battery",              "VARCHAR(100) DEFAULT ''"),
            ("gst_percent",          "FLOAT DEFAULT 0"),
            ("validity_days",        "INTEGER DEFAULT 30"),
            ("vendor_profit",        "FLOAT DEFAULT 0"),
            ("components_json",      "TEXT DEFAULT ''"),
            ("gst_amount",           "FLOAT DEFAULT 0"),
            ("sub_total",            "FLOAT DEFAULT 0"),
            # Pricing snapshot columns
            ("snap_pricing_found",   "BOOLEAN DEFAULT 0"),
            ("snap_vendor_price",    "FLOAT DEFAULT 0"),
            ("snap_hgs_commission",  "FLOAT DEFAULT 0"),
            ("snap_final_price",     "FLOAT DEFAULT 0"),
            ("snap_brand",           "VARCHAR(100) DEFAULT ''"),
            ("snap_warranty_years",  "INTEGER DEFAULT 0"),
            ("snap_panel_price",     "FLOAT DEFAULT 0"),
            ("snap_panel_wattage",   "INTEGER DEFAULT 0"),
            ("snap_panel_quantity",  "INTEGER DEFAULT 0"),
            ("snap_inverter_price",  "FLOAT DEFAULT 0"),
            ("snap_inverter_brand",  "VARCHAR(100) DEFAULT ''"),
            ("snap_inverter_cap",    "VARCHAR(50) DEFAULT ''"),
            ("snap_structure_price", "FLOAT DEFAULT 0"),
            ("snap_structure_type",  "VARCHAR(100) DEFAULT ''"),
            ("snap_install_charge",  "FLOAT DEFAULT 0"),
            ("snap_elec_material",   "FLOAT DEFAULT 0"),
            ("snap_net_meter",       "FLOAT DEFAULT 0"),
            ("snap_transportation",  "FLOAT DEFAULT 0"),
            ("snap_documentation",   "FLOAT DEFAULT 0"),
            ("snap_miscellaneous",   "FLOAT DEFAULT 0"),
            ("snap_gst_percent",     "FLOAT DEFAULT 0"),
        ]
        for col, defn in new_vq_cols:
            if col not in vq2:
                cur2.execute(f"ALTER TABLE vendor_quotation ADD COLUMN {col} {defn}")

        # VendorRating — create if missing
        cur2.execute("""CREATE TABLE IF NOT EXISTS vendor_rating (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            vendor_id  INTEGER NOT NULL,
            lead_id    INTEGER,
            rating     FLOAT DEFAULT 0,
            review     TEXT DEFAULT '',
            rated_by   VARCHAR(100) DEFAULT '',
            created_at VARCHAR(50) DEFAULT '',
            FOREIGN KEY(vendor_id) REFERENCES vendor(id),
            FOREIGN KEY(lead_id)   REFERENCES lead(id)
        )""")

        # Notification — create if missing
        cur2.execute("""CREATE TABLE IF NOT EXISTS notification (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            target_type VARCHAR(20) DEFAULT 'admin',
            target_id   INTEGER,
            title       VARCHAR(200) DEFAULT '',
            message     TEXT DEFAULT '',
            is_read     BOOLEAN DEFAULT 0,
            created_at  VARCHAR(50) DEFAULT ''
        )""")

        conn2.commit()
        conn2.close()

    if not User.query.filter_by(username="admin").first():
        db.session.add(User(
            name="Super Administrator",
            username="admin",
            password=generate_password_hash("admin123"),
            role="admin",
            is_super=True,
            created_at=datetime.now().strftime("%d-%m-%Y")
        ))
        db.session.commit()
    else:
        # Ensure the admin account always has super flag set
        admin_user = User.query.filter_by(username="admin").first()
        if not admin_user.is_super:
            admin_user.is_super = True
            db.session.commit()

    if not VisitorCount.query.first():
        db.session.add(VisitorCount(count=0))
        db.session.commit()

# ==========================
# HELPERS
# ==========================
def add_timeline(lead_id, event, note="", created_by="Admin"):
    tl = LeadTimeline(
        lead_id=lead_id,
        event=event,
        note=note,
        created_by=created_by,
        created_at=datetime.now().strftime("%Y-%m-%d %H:%M")
    )
    db.session.add(tl)
    db.session.commit()

# ==========================
# WEBSITE ROUTES
# ==========================
@app.route("/")
def home():
    # Increment visitor counter on homepage visit
    try:
        vc = VisitorCount.query.first()
        if vc:
            vc.count += 1
            db.session.commit()
    except Exception:
        pass
    return render_template("index.html")

@app.route("/visitor-count")
def visitor_count():
    try:
        vc = VisitorCount.query.first()
        return jsonify({"count": vc.count if vc else 0})
    except Exception:
        return jsonify({"count": 0})

@app.route("/about")
def about():
    return render_template("about.html")

@app.route("/services")
def services():
    return render_template("services.html")

# ==========================
# CONTACT FORM
# ==========================
@app.route("/contact", methods=["GET", "POST"])
def contact():
    if request.method == "POST":
        lead = Lead(
            name=request.form["name"],
            phone=request.form["phone"],
            city=request.form.get("city", ""),
            district=request.form.get("district", ""),
            bill=request.form["bill"],
            created_at=datetime.now().strftime("%Y-%m-%d"),
            status="New"
        )
        db.session.add(lead)
        db.session.commit()
        add_timeline(lead.id, "Lead Created",
                     f"District: {lead.district}", "Customer")
        send_lead_email(lead)
        return render_template("thankyou.html", name=lead.name, city=lead.city)
    return render_template("contact.html", districts=UP_DISTRICTS)

# ==========================
# ADMIN LOGIN
# ==========================
@app.route("/admin-login", methods=["GET", "POST"])
@rate_limit(max_calls=10, period=60)
def admin_login():
    if "user_id" in session:
        return redirect(url_for("admin"))
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "")
        user = User.query.filter_by(username=username).first()
        if user and user.role != "disabled" and check_password_hash(user.password, password):
            session["user_id"]    = user.id
            session["username"]   = user.username
            session["role"]       = user.role
            session["is_super"]   = user.is_super
            session["user_name"]  = user.name or user.username
            session["user_email"] = user.email or ""
            session["last_login"] = datetime.now().strftime("%d-%m-%Y %H:%M")
            logger.info("Admin login: %s", username)
            return redirect(url_for("admin"))
        flash("Invalid credentials.", "danger")
        logger.warning("Failed login attempt for username: %s", username)
    return render_template("login.html")

# ==========================
# ADMIN DASHBOARD
# ==========================
@app.route("/admin")
@login_required
def admin():

    # ── filters ──────────────────────────────────────────────
    f_district = request.args.get("district", "")
    f_vendor   = request.args.get("vendor", "")
    f_status   = request.args.get("status", "")
    f_date     = request.args.get("date", "")
    f_search   = request.args.get("search", "")

    query = Lead.query
    if f_district:
        query = query.filter(Lead.district == f_district)
    if f_vendor:
        query = query.filter(Lead.vendor_id == int(f_vendor))
    if f_status:
        query = query.filter(Lead.status == f_status)
    if f_date:
        query = query.filter(Lead.created_at == f_date)
    if f_search:
        like = f"%{f_search}%"
        query = query.filter(db.or_(
            Lead.name.like(like),
            Lead.phone.like(like),
            Lead.district.like(like)
        ))

    # ── Pagination ────────────────────────────────────────────
    try:
        page     = max(1, int(request.args.get("page", 1)))
        per_page = int(request.args.get("per_page", 25))
        if per_page not in (10, 25, 50, 100):
            per_page = 25
    except (ValueError, TypeError):
        page = 1; per_page = 25

    leads_total_count = query.count()
    leads_total_pages = max(1, (leads_total_count + per_page - 1) // per_page)
    page              = min(page, leads_total_pages)
    leads             = query.order_by(Lead.id.desc()).offset((page - 1) * per_page).limit(per_page).all()

    from flask import url_for as _uf
    leads_base_url = _uf("admin",
        district=f_district, vendor=f_vendor,
        status=f_status, date=f_date, search=f_search)

    users   = User.query.all()
    vendors = Vendor.query.filter_by(is_active=True).all()

    # ── date helpers ─────────────────────────────────────────
    now       = datetime.now()
    today     = now.strftime("%Y-%m-%d")
    yesterday = (now - timedelta(days=1)).strftime("%Y-%m-%d")
    this_month_prefix = now.strftime("%Y-%m")

    # ── stat card counts ─────────────────────────────────────
    total_leads         = Lead.query.count()
    today_leads_count   = Lead.query.filter(Lead.created_at == today).count()
    yesterday_leads_count = Lead.query.filter(Lead.created_at == yesterday).count()
    this_month_count    = Lead.query.filter(
        Lead.created_at.like(f"{this_month_prefix}%")
    ).count()

    status_counts = {s: Lead.query.filter_by(status=s).count() for s in LEAD_STATUSES}

    pending_statuses    = ["New", "Assigned", "Contacted", "Site Visit", "Quotation Sent", "Installation"]
    pending_leads_count = sum(status_counts[s] for s in pending_statuses)
    assigned_leads_count   = status_counts.get("Assigned", 0)
    completed_leads_count  = status_counts.get("Completed", 0)
    cancelled_leads_count  = status_counts.get("Cancelled", 0)

    total_vendors_count    = Vendor.query.count()
    active_vendors_count   = Vendor.query.filter_by(is_active=True).count()
    inactive_vendors_count = total_vendors_count - active_vendors_count

    # ── follow-ups ───────────────────────────────────────────
    today_followups   = Lead.query.filter_by(follow_date=today).all()
    overdue_followups = Lead.query.filter(
        Lead.follow_date < today,
        Lead.follow_date != "",
        Lead.status.notin_(["Completed", "Cancelled"])
    ).all()

    # ── monthly leads chart — last 12 months ─────────────────
    raw_monthly = db.session.query(
        db.func.substr(Lead.created_at, 1, 7).label("ym"),
        db.func.count(Lead.id).label("cnt")
    ).filter(
        Lead.created_at != "",
        Lead.created_at.isnot(None)
    ).group_by("ym").order_by("ym").all()

    monthly_lookup = {row.ym: row.cnt for row in raw_monthly if row.ym}
    month_labels, month_counts = [], []
    for i in range(11, -1, -1):
        d   = now - timedelta(days=i * 30)
        ym  = d.strftime("%Y-%m")
        lbl = d.strftime("%b %Y")
        month_labels.append(lbl)
        month_counts.append(monthly_lookup.get(ym, 0))

    # ── status distribution ───────────────────────────────────
    _status_rows = db.session.query(
        Lead.status, db.func.count(Lead.id)
    ).filter(Lead.status != "", Lead.status != None).group_by(Lead.status).all()
    status_data = [{"label": r[0], "count": r[1]} for r in _status_rows]

    # ── top 10 districts ─────────────────────────────────────
    _district_rows = db.session.query(
        Lead.district, db.func.count(Lead.id).label("cnt")
    ).filter(Lead.district != "", Lead.district != None).group_by(Lead.district).order_by(
        db.func.count(Lead.id).desc()
    ).limit(10).all()
    district_data = [{"label": r[0], "count": r[1]} for r in _district_rows]

    # ── vendor performance (top 10) ───────────────────────────
    all_vendors_list = Vendor.query.all()
    vendor_perf = []
    for v in all_vendors_list:
        v_leads     = Lead.query.filter_by(vendor_id=v.id).count()
        if v_leads == 0:
            continue
        v_completed = Lead.query.filter_by(vendor_id=v.id, status="Completed").count()
        v_pending   = v_leads - v_completed
        v_pct       = round((v_completed / v_leads) * 100) if v_leads else 0
        vendor_perf.append({
            "name":      v.company_name,
            "total":     v_leads,
            "completed": v_completed,
            "pending":   v_pending,
            "pct":       v_pct
        })
    vendor_perf.sort(key=lambda x: x["total"], reverse=True)
    vendor_perf = vendor_perf[:10]

    # ── recent leads (latest 10) ──────────────────────────────
    recent_leads = Lead.query.order_by(Lead.id.desc()).limit(10).all()

    # ── recent activities (latest 10 timeline entries) ────────
    recent_activities = LeadTimeline.query.order_by(
        LeadTimeline.id.desc()
    ).limit(10).all()

    return render_template(
        "admin.html",
        leads=leads, vendors=vendors,
        districts=UP_DISTRICTS, STATUSES=LEAD_STATUSES,
        # stat cards
        status_counts=status_counts,
        total_leads=total_leads,
        today_leads_count=today_leads_count,
        yesterday_leads_count=yesterday_leads_count,
        this_month_count=this_month_count,
        pending_leads_count=pending_leads_count,
        assigned_leads_count=assigned_leads_count,
        completed_leads_count=completed_leads_count,
        cancelled_leads_count=cancelled_leads_count,
        total_vendors_count=total_vendors_count,
        active_vendors_count=active_vendors_count,
        inactive_vendors_count=inactive_vendors_count,
        # charts
        month_labels=month_labels,
        month_counts=month_counts,
        status_data=status_data,
        district_data=district_data,
        vendor_perf=vendor_perf,
        # recent / follow-ups
        recent_leads=recent_leads,
        recent_activities=recent_activities,
        today_followups=today_followups,
        overdue_followups=overdue_followups,
        # filters
        f_district=f_district, f_vendor=f_vendor,
        f_status=f_status, f_date=f_date, f_search=f_search,
        # pagination
        page=page, per_page=per_page,
        leads_total_count=leads_total_count, leads_total_pages=leads_total_pages,
        leads_base_url=leads_base_url
    )

# ==========================
# ADD EMPLOYEE
# ==========================
# ==========================
# ADD ADMIN USER (Super Admin only)
# ==========================
@app.route("/add-user", methods=["POST"])
@super_admin_required
def add_user():
    name     = request.form.get("name", "").strip()
    username = request.form.get("username", "").strip()
    email    = request.form.get("email", "").strip()
    role     = request.form.get("role", "admin")
    password = request.form.get("password", "").strip()

    if not username or not password:
        flash("Username and password are required.", "danger")
        return redirect(url_for("admin"))
    if User.query.filter_by(username=username).first():
        flash("Username already exists.", "danger")
        return redirect(url_for("admin"))

    user = User(
        name=name,
        username=username,
        email=email,
        password=generate_password_hash(password),
        role=role,
        is_super=False,
        created_at=datetime.now().strftime("%d-%m-%Y")
    )
    db.session.add(user)
    db.session.commit()
    flash(f"Admin '{username}' created successfully.", "success")
    return redirect(url_for("admin"))

# ==========================
# CREATE ADMIN — JSON (returns credentials for modal)
# ==========================
@app.route("/create-admin", methods=["POST"])
@super_admin_required
def create_admin():
    import json
    name     = request.form.get("name", "").strip() or "New Admin"
    role     = request.form.get("role", "admin")
    email    = request.form.get("email", "").strip()

    base     = name.lower().replace(" ", ".")[:12]
    suffix   = secrets.token_hex(3)
    username = f"{base}.{suffix}"
    while User.query.filter_by(username=username).first():
        username = f"{base}.{secrets.token_hex(3)}"

    raw_pw = secrets.token_urlsafe(10)
    user   = User(
        name=name,
        username=username,
        email=email,
        password=generate_password_hash(raw_pw),
        role=role,
        is_super=False,
        created_at=datetime.now().strftime("%d-%m-%Y")
    )
    db.session.add(user)
    db.session.commit()
    logger.info("Super admin created new admin user: %s", username)
    return json.dumps({
        "ok":       True,
        "name":     name,
        "username": username,
        "password": raw_pw,
        "email":    email,
        "role":     role
    }), 200, {"Content-Type": "application/json"}

# ==========================
# DISABLE / ENABLE ADMIN USER (Super Admin only)
# ==========================
@app.route("/toggle-user/<int:user_id>")
@super_admin_required
def toggle_user(user_id):
    user = User.query.get_or_404(user_id)
    if user.id == session["user_id"]:
        flash("You cannot disable your own account.", "warning")
        return redirect(url_for("admin"))
    if user.role == "disabled":
        user.role = "admin"
        flash(f"User '{user.username}' enabled.", "success")
    else:
        user.role = "disabled"
        flash(f"User '{user.username}' disabled.", "warning")
    db.session.commit()
    return redirect(url_for("admin"))

# ==========================
# RESET USER PASSWORD (Super Admin only)
# ==========================
@app.route("/reset-password/<int:user_id>", methods=["POST"])
@super_admin_required
def reset_password(user_id):
    user = User.query.get_or_404(user_id)
    new_pw = request.form.get("new_password", "").strip()
    if len(new_pw) < 6:
        flash("Password must be at least 6 characters.", "danger")
        return redirect(url_for("admin"))
    user.password = generate_password_hash(new_pw)
    db.session.commit()
    flash(f"Password reset for '{user.username}'.", "success")
    return redirect(url_for("admin"))

# ==========================
# DELETE ADMIN USER (Super Admin only)
# ==========================
@app.route("/delete-user/<int:user_id>")
@super_admin_required
def delete_user(user_id):
    user = User.query.get_or_404(user_id)
    if user.id == session["user_id"]:
        flash("You cannot delete your own account.", "warning")
        return redirect(url_for("admin"))
    if user.is_super:
        flash("Cannot delete the Super Admin account.", "danger")
        return redirect(url_for("admin"))
    db.session.delete(user)
    db.session.commit()
    flash(f"User '{user.username}' deleted.", "success")
    return redirect(url_for("admin"))

# ==========================
# UPDATE STATUS
# ==========================
@app.route("/update-status/<int:id>/<status>")
def update_status(id, status):
    if "user_id" not in session and "vendor_id" not in session:
        return redirect(url_for("admin_login"))
    if status not in LEAD_STATUSES:
        abort(400)
    lead = Lead.query.get_or_404(id)
    old_status = lead.status
    lead.status = status
    actor = session.get("username") or session.get("vendor_username", "Vendor")
    lead.updated_by = actor
    db.session.commit()
    add_timeline(lead.id, f"Status Changed: {old_status} → {status}", "", actor)
    if "vendor_id" in session:
        return redirect(url_for("vendor_dashboard"))
    return redirect(url_for("admin"))

# ==========================
# ADD NOTE / FOLLOW UP
# ==========================
@app.route("/add-note/<int:id>", methods=["POST"])
def add_note(id):
    if "user_id" not in session and "vendor_id" not in session:
        return redirect(url_for("admin_login"))
    lead = Lead.query.get_or_404(id)
    new_note    = request.form.get("note", "")
    follow_date = request.form.get("follow_date", "")
    lead.note        = new_note
    lead.follow_date = follow_date
    actor            = session.get("username") or session.get("vendor_username", "Vendor")
    lead.updated_by  = actor
    db.session.commit()
    add_timeline(lead.id, "Note Updated", new_note, actor)
    if "vendor_id" in session:
        return redirect(url_for("vendor_dashboard"))
    return redirect(url_for("admin"))

# ==========================
# DELETE LEAD (admin only)
# ==========================
@app.route("/delete/<int:id>")
@admin_required
def delete(id):
    lead = Lead.query.get_or_404(id)
    LeadTimeline.query.filter_by(lead_id=id).delete()
    db.session.delete(lead)
    db.session.commit()
    flash("Lead deleted.", "success")
    return redirect(url_for("admin"))

# ==========================
# ASSIGN VENDOR TO LEAD
# ==========================
@app.route("/assign-vendor/<int:lead_id>", methods=["POST"])
@admin_required
def assign_vendor(lead_id):
    """AJAX vendor assignment — returns JSON, no page redirect."""
    lead      = Lead.query.get_or_404(lead_id)
    vendor_id = request.form.get("vendor_id") or (
        request.get_json(silent=True) or {}
    ).get("vendor_id")

    if not vendor_id:
        return jsonify({"ok": False, "error": "No vendor selected."}), 400

    try:
        vendor_id = int(vendor_id)
    except (ValueError, TypeError):
        return jsonify({"ok": False, "error": "Invalid vendor ID."}), 400

    vendor = Vendor.query.get(vendor_id)
    if not vendor:
        return jsonify({"ok": False, "error": "Vendor not found."}), 404

    prev_vendor_name = lead.vendor.company_name if lead.vendor else None
    force            = (request.form.get("force") == "1") or (
        (request.get_json(silent=True) or {}).get("force") == True
    )

    # Warn on duplicate assignment (unless forced)
    if lead.vendor_id and lead.vendor_id != vendor_id and not force:
        return jsonify({
            "ok": False,
            "duplicate": True,
            "prev_vendor": prev_vendor_name,
            "message": f"This lead is already assigned to {prev_vendor_name}. Replace?"
        }), 200

    lead.vendor_id  = vendor_id
    lead.status     = "Assigned"
    lead.updated_by = session.get("username", "Admin")
    db.session.commit()

    admin_name = session.get("user_name") or session.get("username", "Admin")
    add_timeline(
        lead_id,
        f"Vendor Assigned: {vendor.company_name}",
        f"Assigned by {admin_name}" + (f" (replaced {prev_vendor_name})" if prev_vendor_name and prev_vendor_name != vendor.company_name else ""),
        admin_name
    )

    return jsonify({
        "ok": True,
        "vendor_id":   vendor.id,
        "vendor_name": vendor.company_name,
        "district":    vendor.district,
        "message":     f"Vendor '{vendor.company_name}' assigned successfully."
    })


@app.route("/vendor-search")
@login_required
def vendor_search():
    """AJAX searchable vendor endpoint for Tom Select with district filter + scoring."""
    import json as _json, statistics
    q          = request.args.get("q", "").strip()
    lead_id    = request.args.get("lead_id", type=int)
    district   = request.args.get("district", "").strip()
    page       = request.args.get("page", 1, type=int)
    per_page   = 20

    # Base query — active vendors only
    qry = Vendor.query.filter_by(is_active=True)

    # District filter — intersect with coverage
    if district:
        # Collect vendor IDs that cover this district
        cov_all     = VendorCoverage.query.filter_by(all_districts=True).with_entities(VendorCoverage.vendor_id).all()
        cov_all_ids = {r[0] for r in cov_all}
        cov_dist    = VendorCoverage.query.filter(
            VendorCoverage.all_districts == False
        ).all()
        cov_dist_ids = set()
        for c in cov_dist:
            try:
                dl = _json.loads(c.districts or "[]")
            except Exception:
                dl = []
            if district in dl:
                cov_dist_ids.add(c.vendor_id)
        # Also include vendors with no coverage record but matching base district
        no_cov_ids = {v.id for v in Vendor.query.filter_by(
            is_active=True, district=district
        ).filter(
            ~Vendor.id.in_(
                db.session.query(VendorCoverage.vendor_id).subquery()
            )
        ).all()}
        allowed_ids = cov_all_ids | cov_dist_ids | no_cov_ids
        if allowed_ids:
            qry = qry.filter(Vendor.id.in_(allowed_ids))
        else:
            return jsonify({"results": [], "has_more": False})

    # Text search
    if q:
        like = f"%{q}%"
        qry = qry.filter(db.or_(
            Vendor.company_name.ilike(like),
            Vendor.owner_name.ilike(like),
            Vendor.district.ilike(like),
            Vendor.mobile.ilike(like),
            Vendor.username.ilike(like),
        ))

    total   = qry.count()
    vendors = qry.offset((page - 1) * per_page).limit(per_page).all()

    results = []
    for v in vendors:
        ratings   = [r.rating for r in VendorRating.query.filter_by(vendor_id=v.id).all()]
        avg_r     = round(statistics.mean(ratings), 1) if ratings else 0.0
        total_l   = Lead.query.filter_by(vendor_id=v.id).count()
        done_l    = Lead.query.filter_by(vendor_id=v.id, status="Completed").count()
        conv      = round(done_l / total_l * 100, 1) if total_l else 0.0
        profile   = VendorProfile.query.filter_by(vendor_id=v.id).first()
        yrs_exp   = profile.years_experience if profile else 0

        # Scoring for sort/recommendation
        score = (avg_r * 20) + (conv * 0.5) + (min(done_l, 200) * 0.1)
        stars = "★" * int(avg_r) + ("½" if avg_r % 1 >= 0.5 else "") + "☆" * (5 - int(avg_r) - (1 if avg_r % 1 >= 0.5 else 0))

        results.append({
            "id":           v.id,
            "text":         f"{v.company_name} — {v.district}",
            "company_name": v.company_name,
            "owner_name":   v.owner_name,
            "district":     v.district,
            "mobile":       v.mobile,
            "avg_rating":   avg_r,
            "stars":        stars,
            "done":         done_l,
            "conv":         conv,
            "yrs_exp":      yrs_exp,
            "score":        score,
        })

    # Sort by score descending; mark top as recommended
    results.sort(key=lambda x: -x["score"])
    if results:
        results[0]["recommended"] = True

    return jsonify({"results": results, "has_more": (page * per_page) < total})


@app.route("/eligible-vendors/<int:lead_id>")
@login_required
def eligible_vendors(lead_id):
    """Return vendors eligible for a lead based on district & lead type coverage."""
    import json
    lead = Lead.query.get_or_404(lead_id)
    active_vendors = Vendor.query.filter_by(is_active=True).all()
    eligible = []
    for v in active_vendors:
        cov = VendorCoverage.query.filter_by(vendor_id=v.id).first()
        if cov:
            if cov.all_districts:
                eligible.append(v)
                continue
            try:
                dist_list = json.loads(cov.districts or "[]")
            except Exception:
                dist_list = []
            if lead.district and lead.district in dist_list:
                eligible.append(v)
        else:
            # No coverage set — include all vendors that match base district
            if not lead.district or v.district == lead.district:
                eligible.append(v)
    return jsonify([{
        "id": v.id,
        "company_name": v.company_name,
        "district": v.district,
        "mobile": v.mobile
    } for v in eligible])

# ==========================
# LEAD TIMELINE API
# ==========================
@app.route("/lead-timeline/<int:lead_id>")
def lead_timeline(lead_id):
    if "user_id" not in session:
        return jsonify([])
    entries = LeadTimeline.query.filter_by(lead_id=lead_id).order_by(
        LeadTimeline.id.asc()
    ).all()
    return jsonify([{
        "event": e.event,
        "note": e.note,
        "created_by": e.created_by,
        "created_at": e.created_at
    } for e in entries])

# ==========================
# VENDOR LOGIN
# ==========================
@app.route("/vendor-login", methods=["GET", "POST"])
@rate_limit(max_calls=10, period=60)
def vendor_login():
    if request.method == "POST":
        username = request.form["username"]
        password = request.form["password"]
        vendor = Vendor.query.filter_by(username=username, is_active=True).first()
        if vendor and check_password_hash(vendor.password, password):
            session["vendor_id"]       = vendor.id
            session["vendor_username"] = vendor.username
            session["vendor_district"] = vendor.district
            session["vendor_company"]  = vendor.company_name
            session["vendor_name"]     = vendor.owner_name
            session["vendor_mobile"]   = vendor.mobile
            session["vendor_email"]    = vendor.email or ""
            return redirect("/vendor-dashboard")
        flash("Invalid credentials or inactive account.", "danger")
    return render_template("vendor_login.html")

# ==========================
# VENDOR LOGOUT
# ==========================
@app.route("/vendor-logout")
def vendor_logout():
    session.pop("vendor_id", None)
    session.pop("vendor_username", None)
    session.pop("vendor_district", None)
    session.pop("vendor_company", None)
    session.pop("vendor_name", None)
    session.pop("vendor_mobile", None)
    session.pop("vendor_email", None)
    return redirect("/vendor-login")

# ==========================
# VENDOR PROFILE
# ==========================
@app.route("/vendor-profile", methods=["GET", "POST"])
@vendor_required
def vendor_profile():
    vendor = Vendor.query.get_or_404(session["vendor_id"])
    profile = VendorProfile.query.filter_by(vendor_id=vendor.id).first()
    if not profile:
        profile = VendorProfile(vendor_id=vendor.id)
        db.session.add(profile)
        db.session.commit()
    if request.method == "POST":
        # Update Vendor base fields
        vendor.company_name  = request.form.get("company_name", vendor.company_name).strip()
        vendor.owner_name    = request.form.get("owner_name", vendor.owner_name).strip()
        vendor.mobile        = request.form.get("mobile", vendor.mobile).strip()
        vendor.email         = request.form.get("email", vendor.email).strip()
        vendor.address       = request.form.get("address", vendor.address).strip()
        vendor.district      = request.form.get("district", vendor.district).strip()
        # Update profile fields
        profile.gst_number        = request.form.get("gst_number", "").strip()
        profile.pan_number        = request.form.get("pan_number", "").strip()
        profile.state             = request.form.get("state", "Uttar Pradesh").strip()
        profile.pincode           = request.form.get("pincode", "").strip()
        profile.alternate_phone   = request.form.get("alternate_phone", "").strip()
        profile.website           = request.form.get("website", "").strip()
        try:
            profile.years_experience   = int(request.form.get("years_experience", 0))
            profile.install_capacity_kw= int(request.form.get("install_capacity_kw", 0))
            profile.res_projects       = int(request.form.get("res_projects", 0))
            profile.comm_projects      = int(request.form.get("comm_projects", 0))
            profile.ind_projects       = int(request.form.get("ind_projects", 0))
            profile.govt_projects      = int(request.form.get("govt_projects", 0))
        except (ValueError, TypeError):
            pass
        # BUG 1 FIX: hidden(value=0) + checkbox(value=1) both submit when checked.
        # request.form.get() returns the FIRST value ("0"). Must use getlist and check for "1".
        profile.pm_surya_approved  = "1" in request.form.getlist("pm_surya_approved")
        profile.discom_empanelled  = "1" in request.form.getlist("discom_empanelled")
        profile.brands_supported   = ",".join(request.form.getlist("brands_supported"))
        # Bank / service fields (BUG 6: were missing from save)
        profile.bank_name          = request.form.get("bank_name", "").strip()
        profile.account_holder     = request.form.get("account_holder", "").strip()
        profile.account_number     = request.form.get("account_number", "").strip()
        profile.ifsc_code          = request.form.get("ifsc_code", "").strip()
        profile.upi_id             = request.form.get("upi_id", "").strip()
        try:
            profile.service_radius = int(request.form.get("service_radius", 0) or 0)
        except (ValueError, TypeError):
            profile.service_radius = 0
        profile.updated_at         = datetime.now().strftime("%d-%m-%Y %H:%M")
        db.session.commit()
        # Update session company name
        session["vendor_company"] = vendor.company_name
        session["vendor_name"]    = vendor.owner_name
        return jsonify({"ok": True, "message": "Profile updated successfully."})
    return render_template("vendor_profile.html",
        vendor=vendor, profile=profile,
        districts=UP_DISTRICTS, brands=SOLAR_BRANDS)


# ==========================
# VENDOR PRICING
# ==========================
@app.route("/vendor-pricing")
@vendor_required
def vendor_pricing():
    vendor = Vendor.query.get_or_404(session["vendor_id"])
    pricing = VendorPricing.query.filter_by(vendor_id=vendor.id).all()
    pricing_map = {(p.project_type, p.capacity_kw): p for p in pricing}
    return render_template("vendor_pricing.html",
        vendor=vendor, pricing_map=pricing_map,
        capacities=KW_CAPACITIES, brands=SOLAR_BRANDS)


@app.route("/vendor-pricing/save", methods=["POST"])
@vendor_required
def vendor_pricing_save():
    vendor_id = session["vendor_id"]
    now  = datetime.now().strftime("%d-%m-%Y %H:%M")
    data = request.get_json()
    if not data:
        return jsonify({"ok": False, "error": "No data received."}), 400

    saved_records = []

    for row in data:
        # ── Capacity ──
        raw_cap = str(row.get("capacity_kw", "")).strip()
        cap = int(raw_cap) if raw_cap.isdigit() and int(raw_cap) > 0 else 0
        if not cap:
            continue

        ptype = row.get("ptype", "Residential").strip()
        if ptype not in ("Residential", "Commercial", "Government"):
            continue

        # ── Safe float/int helpers: ignore empty strings, keep existing value otherwise ──
        def _f(key, existing_val=0.0):
            v = row.get(key)
            if v is None or str(v).strip() == "":
                return existing_val  # do not overwrite with empty
            try:
                result = float(v)
                return result if result >= 0 else existing_val
            except (ValueError, TypeError):
                return existing_val

        def _i(key, existing_val=0):
            v = row.get(key)
            if v is None or str(v).strip() == "":
                return existing_val
            try:
                result = int(float(v))
                return result if result >= 0 else existing_val
            except (ValueError, TypeError):
                return existing_val

        def _s(key, existing_val=""):
            v = row.get(key)
            if v is None:
                return existing_val
            stripped = str(v).strip()
            return stripped if stripped else existing_val

        # ── Upsert: find or create — never duplicate ──
        rec = VendorPricing.query.filter_by(
            vendor_id=vendor_id, project_type=ptype, capacity_kw=cap
        ).first()
        is_new = rec is None
        if is_new:
            rec = VendorPricing(vendor_id=vendor_id, project_type=ptype, capacity_kw=cap)
            db.session.add(rec)
            db.session.flush()  # get rec populated before reading existing vals

        # ── Validate selling price (must be present and positive for the row to be saved) ──
        vp_raw = row.get("vendor_selling_price")
        if vp_raw is None or str(vp_raw).strip() == "":
            if is_new:
                db.session.expunge(rec)
                continue
            # For existing record keep current value — skip overwrite
        else:
            try:
                vp = float(vp_raw)
                if vp < 0:
                    return jsonify({"ok": False, "error": f"Vendor selling price cannot be negative for {cap} KW {ptype}."}), 400
            except (ValueError, TypeError):
                return jsonify({"ok": False, "error": f"Invalid selling price for {cap} KW {ptype}."}), 400

        hgs_raw = row.get("hgs_commission")
        hgs = 0.0
        if hgs_raw is not None and str(hgs_raw).strip() != "":
            try:
                hgs = float(hgs_raw)
                if hgs < 0:
                    return jsonify({"ok": False, "error": f"Commission cannot be negative for {cap} KW {ptype}."}), 400
            except (ValueError, TypeError):
                hgs = rec.hgs_commission or 0.0

        vp_final = _f("vendor_selling_price", rec.vendor_price or 0.0)
        if hgs > vp_final and vp_final > 0:
            return jsonify({"ok": False,
                "error": f"Commission (₹{hgs:,.0f}) cannot exceed selling price (₹{vp_final:,.0f}) for {cap} KW {ptype}."}), 400

        # ── Apply fields: only overwrite when new value is provided ──
        rec.vendor_price      = vp_final
        rec.hgs_commission    = _f("hgs_commission",      rec.hgs_commission    or 0.0)
        rec.final_price       = round(rec.vendor_price + rec.hgs_commission, 2)
        rec.brand             = _s("brand",               rec.brand             or "")
        rec.warranty_years    = _i("warranty_years",      rec.warranty_years    or 0)
        rec.system_price      = rec.vendor_price  # keep in sync

        # Component fields — only overwrite if provided
        rec.panel_price       = _f("panel_price",         rec.panel_price       or 0.0)
        rec.panel_wattage     = _i("panel_wattage",       rec.panel_wattage     or 0)
        rec.panel_quantity    = _i("panel_quantity",      rec.panel_quantity    or 0)
        rec.inverter_price    = _f("inverter_price",      rec.inverter_price    or 0.0)
        rec.inverter_brand    = _s("inverter_brand",      rec.inverter_brand    or "")
        rec.inverter_capacity = _s("inverter_capacity",   rec.inverter_capacity or "")
        rec.structure_price   = _f("structure_price",     rec.structure_price   or 0.0)
        rec.structure_type    = _s("structure_type",      rec.structure_type    or "")
        rec.install_charge    = _f("install_charge",      rec.install_charge    or 0.0)
        rec.elec_material     = _f("elec_material",       rec.elec_material     or 0.0)
        rec.net_meter_charge  = _f("net_meter_charge",    rec.net_meter_charge  or 0.0)
        rec.transportation    = _f("transportation",      rec.transportation    or 0.0)
        rec.documentation     = _f("documentation",       rec.documentation     or 0.0)
        rec.miscellaneous     = _f("miscellaneous",       rec.miscellaneous     or 0.0)
        rec.gst_percent       = _f("gst_percent",         rec.gst_percent       or 0.0)
        rec.updated_at        = now

    # ── Single commit for all rows ──
    try:
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        logger.error("Pricing save error: %s", e)
        return jsonify({"ok": False, "error": "Database error. Please try again."}), 500

    # ── Re-read from DB and return all saved values so JS can reload the card ──
    cap_to_return = None
    if data:
        raw_cap = str(data[0].get("capacity_kw", "")).strip()
        cap_to_return = int(raw_cap) if raw_cap.isdigit() else None

    result_map = {}
    if cap_to_return:
        records = VendorPricing.query.filter_by(vendor_id=vendor_id, capacity_kw=cap_to_return).all()
        for r in records:
            result_map[r.project_type] = {
                "brand":            r.brand or "",
                "warranty_years":   r.warranty_years or 0,
                "vendor_price":     r.vendor_price or 0,
                "hgs_commission":   r.hgs_commission or 0,
                "final_price":      r.final_price or 0,
                "panel_price":      r.panel_price or 0,
                "panel_wattage":    r.panel_wattage or 0,
                "panel_quantity":   r.panel_quantity or 0,
                "inverter_price":   r.inverter_price or 0,
                "inverter_brand":   r.inverter_brand or "",
                "inverter_capacity":r.inverter_capacity or "",
                "structure_price":  r.structure_price or 0,
                "structure_type":   r.structure_type or "",
                "install_charge":   r.install_charge or 0,
                "elec_material":    r.elec_material or 0,
                "net_meter_charge": r.net_meter_charge or 0,
                "transportation":   r.transportation or 0,
                "documentation":    r.documentation or 0,
                "miscellaneous":    r.miscellaneous or 0,
                "gst_percent":      r.gst_percent or 0,
                "updated_at":       r.updated_at or "",
            }

    return jsonify({"ok": True, "message": "Pricing saved successfully.", "saved": result_map})


# ==========================
# QUOTATION REVISION WORKFLOW
# ==========================
@app.route("/vendor-quotation/<int:qid>/vendor-action", methods=["POST"])
@vendor_required
def vendor_quotation_vendor_action(qid):
    """Vendor acts on a Revision request: Accept / Reject / Edit+Resubmit."""
    q = VendorQuotation.query.get_or_404(qid)
    if q.vendor_id != session["vendor_id"]:
        return jsonify({"ok": False, "error": "Unauthorized."}), 403
    action = request.form.get("action", "").strip()
    now    = datetime.now().strftime("%d-%m-%Y %H:%M")

    if action == "accept":
        q.status     = "Accepted by Vendor"
        q.updated_at = now
        db.session.commit()
        push_notification("admin", "Vendor Accepted Quotation",
            f"{q.vendor.company_name} accepted quotation #{q.id} at ₹{q.net_price:,.0f}.")
        return jsonify({"ok": True, "status": q.status,
                        "message": "Quotation accepted."})

    elif action == "reject":
        reason = request.form.get("reason", "").strip()
        q.status       = "Rejected by Vendor"
        q.admin_remarks= q.admin_remarks  # keep admin remarks
        q.remarks      = (q.remarks or "") + f"\n[Vendor Rejection] {reason}" if reason else q.remarks
        q.updated_at   = now
        db.session.commit()
        push_notification("admin", "Vendor Rejected Quotation",
            f"{q.vendor.company_name} rejected quotation #{q.id}. Reason: {reason or 'No reason given'}.")
        return jsonify({"ok": True, "status": q.status,
                        "message": "Quotation rejected."})

    elif action == "resubmit":
        # Vendor edits and resubmits revised quotation
        def _f(key):
            try: return float(request.form.get(key) or 0)
            except (ValueError, TypeError): return 0.0
        q.gross_price    = _f("gross_price")
        q.subsidy_amount = _f("subsidy_amount")
        q.net_price      = _f("net_price")
        q.commission     = _f("commission")
        q.remarks        = request.form.get("remarks", q.remarks).strip()
        q.brand          = request.form.get("brand", q.brand).strip()
        q.panel_brand    = request.form.get("panel_brand", q.panel_brand).strip()
        q.inverter_brand = request.form.get("inverter_brand", q.inverter_brand).strip()
        q.status         = "Revision Submitted"
        q.updated_at     = now
        db.session.commit()
        push_notification("admin", "Vendor Submitted Revised Quotation",
            f"{q.vendor.company_name} resubmitted quotation #{q.id} at ₹{q.net_price:,.0f}.")
        push_notification("vendor", "Your Revision Was Submitted",
            f"Your revised quotation #{q.id} has been sent to admin for review.",
            target_id=q.vendor_id)
        return jsonify({"ok": True, "status": q.status,
                        "message": "Revised quotation submitted."})

    return jsonify({"ok": False, "error": "Unknown action."}), 400


@app.route("/admin/quotation/<int:qid>/action", methods=["POST"])
@admin_required
def admin_quotation_action(qid):
    q = VendorQuotation.query.get_or_404(qid)
    action = request.form.get("action")
    now    = datetime.now().strftime("%d-%m-%Y %H:%M")
    if action in ("Approved", "Rejected", "Revision"):
        q.status        = action
        q.admin_remarks = request.form.get("admin_remarks", "").strip()
        q.updated_at    = now
        db.session.commit()
        # Notifications to vendor
        if action == "Approved":
            push_notification("vendor", "Quotation Approved ✓",
                f"Your quotation #{q.id} for ₹{q.net_price:,.0f} has been approved!",
                target_id=q.vendor_id)
        elif action == "Rejected":
            push_notification("vendor", "Quotation Rejected",
                f"Your quotation #{q.id} has been rejected. Remarks: {q.admin_remarks or 'None'}.",
                target_id=q.vendor_id)
        elif action == "Revision":
            push_notification("vendor", "Revision Requested on Quotation",
                f"Admin requested a revision on quotation #{q.id}. Remarks: {q.admin_remarks or 'None'}.",
                target_id=q.vendor_id)
    return jsonify({"ok": True, "status": q.status})
@app.route("/vendor-commission")
@vendor_required
def vendor_commission():
    vendor = Vendor.query.get_or_404(session["vendor_id"])
    comms  = VendorCommission.query.filter_by(vendor_id=vendor.id).all()
    comm_map = {(c.project_type, c.capacity_kw): c for c in comms}
    return render_template("vendor_commission.html",
        vendor=vendor, comm_map=comm_map,
        capacities=KW_CAPACITIES)


@app.route("/vendor-commission/save", methods=["POST"])
@vendor_required
def vendor_commission_save():
    vendor_id = session["vendor_id"]
    now  = datetime.now().strftime("%d-%m-%Y %H:%M")
    data = request.get_json()
    if not data:
        return jsonify({"ok": False, "error": "No data."}), 400
    for row in data:
        ptype = row.get("project_type", "").strip()
        cap   = int(row.get("capacity_kw", 0))
        if not ptype or not cap:
            continue
        existing = VendorCommission.query.filter_by(
            vendor_id=vendor_id, project_type=ptype, capacity_kw=cap
        ).first()
        if not existing:
            existing = VendorCommission(vendor_id=vendor_id, project_type=ptype, capacity_kw=cap)
            db.session.add(existing)
        try:
            vp  = float(row.get("vendor_price") or 0)
            hgs = float(row.get("hgs_commission") or 0)
        except (ValueError, TypeError):
            vp = hgs = 0
        existing.vendor_price   = vp
        existing.hgs_commission = hgs
        existing.final_price    = vp + hgs
        existing.updated_at     = now
    db.session.commit()
    return jsonify({"ok": True, "message": "Commission settings saved."})


# ==========================
# VENDOR COVERAGE
# ==========================
@app.route("/vendor-coverage", methods=["GET", "POST"])
@vendor_required
def vendor_coverage():
    vendor   = Vendor.query.get_or_404(session["vendor_id"])
    coverage = VendorCoverage.query.filter_by(vendor_id=vendor.id).first()
    if not coverage:
        coverage = VendorCoverage(vendor_id=vendor.id)
        db.session.add(coverage)
        db.session.commit()
    if request.method == "POST":
        import json
        selected = request.form.getlist("districts")
        all_dist = request.form.get("all_districts") == "1"
        lead_types = request.form.getlist("lead_types")
        coverage.districts    = json.dumps(selected)
        coverage.all_districts= all_dist
        coverage.lead_types   = json.dumps(lead_types)
        coverage.updated_at   = datetime.now().strftime("%d-%m-%Y %H:%M")
        db.session.commit()
        return jsonify({"ok": True, "message": "Coverage saved."})
    import json
    selected_districts = []
    selected_types     = []
    try:
        selected_districts = json.loads(coverage.districts or "[]")
        selected_types     = json.loads(coverage.lead_types or "[]")
    except Exception:
        pass
    return render_template("vendor_coverage.html",
        vendor=vendor, coverage=coverage,
        all_districts=UP_DISTRICTS, lead_types=LEAD_TYPES,
        selected_districts=selected_districts,
        selected_types=selected_types)


# ==========================
# VENDOR QUOTATION
# ==========================
@app.route("/vendor-quotations")
@vendor_required
def vendor_quotations():
    vendor = Vendor.query.get_or_404(session["vendor_id"])
    quotes = VendorQuotation.query.filter_by(vendor_id=vendor.id)\
                .order_by(VendorQuotation.id.desc()).all()
    return render_template("vendor_quotation.html",
        vendor=vendor, quotes=quotes,
        capacities=KW_CAPACITIES, brands=SOLAR_BRANDS)


@app.route("/vendor-quotation/create", methods=["POST"])
@vendor_required
def vendor_quotation_create():
    import json as _json
    vendor_id = session["vendor_id"]
    now = datetime.now().strftime("%d-%m-%Y %H:%M")

    # ── Parse form inputs ──────────────────────────────────────────────────
    try:
        cap    = int(request.form.get("capacity_kw") or 0)
    except (ValueError, TypeError):
        cap = 0
    try:
        subsidy = float(request.form.get("subsidy_amount") or 0)
    except (ValueError, TypeError):
        subsidy = 0.0
    try:
        # commission is sent from JS as _qPricing.hgs_commission — trust it directly
        comm = float(request.form.get("commission") or 0)
    except (ValueError, TypeError):
        comm = 0.0
    # form_gross comes from vendor_pricing_get calculation in JS (_qPricing.gross_price)
    # It is the AUTHORITATIVE value when the backend's component sum is unreliable
    try:
        form_gross = float(request.form.get("gross_price") or 0)
    except (ValueError, TypeError):
        form_gross = 0.0

    ptype = (request.form.get("project_type") or "Residential").strip()
    if ptype not in ("Residential", "Commercial", "Government"):
        ptype = "Residential"

    # ── Load VendorPricing row — 3-level fallback ──────────────────────────
    pricing = VendorPricing.query.filter_by(
        vendor_id=vendor_id, project_type=ptype, capacity_kw=cap
    ).first()
    if not pricing and cap:
        pricing = VendorPricing.query.filter_by(
            vendor_id=vendor_id, project_type=ptype
        ).order_by(db.func.abs(VendorPricing.capacity_kw - cap)).first()
    if not pricing:
        pricing = VendorPricing.query.filter_by(vendor_id=vendor_id).first()

    pricing_found = pricing is not None
    p = pricing  # alias — safe to use from here on (may be None)

    # ── Build itemised component list ─────────────────────────────────────
    components = []
    sub_total  = 0.0
    gst_pct    = 0.0
    gst_amt    = 0.0
    gross      = 0.0

    if pricing_found:
        panel_qty   = (p.panel_quantity or 0) if p.panel_quantity else max(1, cap)
        panel_price = p.panel_price or 0.0
        panel_total = round(panel_price * panel_qty, 2)

        if panel_total > 0:
            components.append({
                "item":       "Solar Panel",
                "details":    f"{p.brand or '—'} | {p.panel_wattage or '—'} W × {panel_qty}",
                "unit_price": panel_price,
                "qty":        panel_qty,
                "amount":     panel_total,
            })
        if (p.inverter_price or 0) > 0:
            components.append({
                "item":       "Inverter",
                "details":    f"{p.inverter_brand or '—'} | {p.inverter_capacity or f'{cap} KW'}",
                "unit_price": p.inverter_price,
                "qty":        1,
                "amount":     p.inverter_price,
            })
        if (p.structure_price or 0) > 0:
            components.append({
                "item":       "Structure",
                "details":    p.structure_type or "—",
                "unit_price": p.structure_price,
                "qty":        1,
                "amount":     p.structure_price,
            })
        for label, field in [
            ("Installation Charge", "install_charge"),
            ("Electrical Material", "elec_material"),
            ("Net Meter Charge",    "net_meter_charge"),
            ("Transportation",      "transportation"),
            ("Documentation",       "documentation"),
            ("Miscellaneous",       "miscellaneous"),
        ]:
            val = float(getattr(p, field, 0) or 0)
            if val > 0:
                components.append({
                    "item": label, "details": "—",
                    "unit_price": val, "qty": 1, "amount": val,
                })

        # ── Gross from itemised sum ────────────────────────────────────────
        sub_total_comps = round(sum(c["amount"] for c in components), 2)
        gst_pct         = float(p.gst_percent or 0)
        gst_amt_comps   = round(sub_total_comps * gst_pct / 100, 2)
        gross_comps     = round(sub_total_comps + gst_amt_comps, 2)

        if gross_comps > 0:
            # Full itemised breakdown is available — use it
            sub_total = sub_total_comps
            gst_amt   = gst_amt_comps
            gross     = gross_comps
        else:
            # Component fields not filled — use vendor_price as authoritative gross
            # (vendor_price is set by the vendor in the pricing form)
            vp = float(p.vendor_price or 0)
            if vp > 0:
                gst_amt   = round(vp * gst_pct / 100, 2)
                gross     = round(vp + gst_amt, 2)
                sub_total = vp
                components = [{
                    "item":       f"{cap} KW Solar System",
                    "details":    f"{p.brand or ptype} | {p.warranty_years or 0} yr warranty",
                    "unit_price": vp,
                    "qty":        1,
                    "amount":     vp,
                }]
            else:
                # vendor_price also not set — last resort: trust the form value
                # (form_gross was calculated by vendor_pricing_get on the frontend)
                gross     = form_gross
                sub_total = form_gross
                gst_pct   = 0.0
                gst_amt   = 0.0

        # ── Final safety: if gross is still 0 but form sent a valid value, use it ──
        if gross == 0 and form_gross > 0:
            gross     = form_gross
            sub_total = form_gross
            gst_pct   = float(p.gst_percent or 0) if p else 0.0
            gst_amt   = round(gross * gst_pct / 100, 2)
    else:
        # No pricing record at all — use manually entered gross from form
        gross     = form_gross
        sub_total = form_gross
        gst_pct   = 0.0
        gst_amt   = 0.0

    # ── Derived totals ────────────────────────────────────────────────────
    gross   = round(gross, 2)
    net     = round(gross - subsidy + comm, 2)

    # ── Build VendorQuotation row ──────────────────────────────────────────
    q = VendorQuotation(
        vendor_id       = vendor_id,
        lead_id         = request.form.get("lead_id") or None,
        customer_name   = request.form.get("customer_name", "").strip(),
        customer_phone  = request.form.get("customer_phone", "").strip(),
        capacity_kw     = cap,
        project_type    = ptype,
        brand           = (p.brand or "") if p else request.form.get("brand", "").strip(),
        panel_brand     = (p.brand or "") if p else request.form.get("panel_brand", "").strip(),
        inverter_brand  = (p.inverter_brand or "") if p else request.form.get("inverter_brand", "").strip(),
        structure_type  = (p.structure_type or "") if p else request.form.get("structure_type", "").strip(),
        subsidy_amount  = subsidy,
        gross_price     = gross,
        net_price       = net,
        commission      = comm,
        remarks         = request.form.get("remarks", "").strip(),
        status          = "Pending",
        components_json = _json.dumps(components),
        sub_total       = sub_total,
        gst_amount      = gst_amt,
        gst_percent     = gst_pct,
        created_at      = now,
        updated_at      = now,
        # ── Full pricing snapshot — frozen at creation, never changes ──────
        snap_pricing_found   = pricing_found,
        snap_vendor_price    = float(p.vendor_price    or 0) if p else 0.0,
        snap_hgs_commission  = float(p.hgs_commission  or 0) if p else 0.0,
        snap_final_price     = float(p.final_price     or 0) if p else 0.0,
        snap_brand           = (p.brand           or "") if p else "",
        snap_warranty_years  = int(p.warranty_years   or 0) if p else 0,
        snap_panel_price     = float(p.panel_price    or 0) if p else 0.0,
        snap_panel_wattage   = int(p.panel_wattage    or 0) if p else 0,
        snap_panel_quantity  = int(p.panel_quantity   or 0) if p else 0,
        snap_inverter_price  = float(p.inverter_price or 0) if p else 0.0,
        snap_inverter_brand  = (p.inverter_brand  or "") if p else "",
        snap_inverter_cap    = (p.inverter_capacity or "") if p else "",
        snap_structure_price = float(p.structure_price or 0) if p else 0.0,
        snap_structure_type  = (p.structure_type  or "") if p else "",
        snap_install_charge  = float(p.install_charge  or 0) if p else 0.0,
        snap_elec_material   = float(p.elec_material   or 0) if p else 0.0,
        snap_net_meter       = float(p.net_meter_charge or 0) if p else 0.0,
        snap_transportation  = float(p.transportation  or 0) if p else 0.0,
        snap_documentation   = float(p.documentation  or 0) if p else 0.0,
        snap_miscellaneous   = float(p.miscellaneous   or 0) if p else 0.0,
        snap_gst_percent     = float(p.gst_percent     or 0) if p else 0.0,
    )
    db.session.add(q)
    db.session.commit()

    return jsonify({
        "ok":              True,
        "id":              q.id,
        "message":         f"Quotation #{q.id} created. Gross: ₹{gross:,.0f}",
        "gross_price":     gross,
        "pricing_missing": not pricing_found,
    })


@app.route("/vendor-quotation/<int:qid>")
@vendor_required
def vendor_quotation_detail(qid):
    q = VendorQuotation.query.get_or_404(qid)
    if q.vendor_id != session["vendor_id"]:
        abort(403)
    vendor = Vendor.query.get_or_404(session["vendor_id"])

    # ── Recalc gross from snapshot if it was saved as 0 ──────────────────
    # This repairs quotations created before the gross-price fix.
    if (not q.gross_price or q.gross_price == 0):
        recalc = 0.0
        # 1) Try summing component breakdown from components_json
        try:
            import json as _j
            comps = _j.loads(q.components_json or "[]")
            if comps:
                recalc = sum(float(c.get("amount", 0)) for c in comps)
        except Exception:
            recalc = 0.0
        # 2) Fall back to snap_vendor_price (always reliable)
        if recalc == 0 and (q.snap_vendor_price or 0) > 0:
            recalc = float(q.snap_vendor_price)
        # 3) Apply GST on top of recalc base
        if recalc > 0:
            gst_pct = float(q.snap_gst_percent or q.gst_percent or 0)
            gst_on  = round(recalc * gst_pct / 100, 2)
            q.gross_price = round(recalc + gst_on, 2)
            if not q.sub_total or q.sub_total == 0:
                q.sub_total  = recalc
                q.gst_amount = gst_on
            q.net_price   = round(q.gross_price - (q.subsidy_amount or 0) + (q.commission or 0), 2)
            db.session.commit()

    return render_template("vendor_quotation_detail.html", q=q, vendor=vendor)


# ==========================
# ADMIN: QUOTATION APPROVAL
# ==========================
@app.route("/admin/quotations")
@login_required
def admin_quotations():
    try:
        page     = max(1, int(request.args.get("page", 1)))
        per_page = int(request.args.get("per_page", 25))
        if per_page not in (10, 25, 50, 100):
            per_page = 25
    except (ValueError, TypeError):
        page = 1; per_page = 25

    f_status = request.args.get("status", "")
    f_vendor = request.args.get("vendor_id", "")
    query    = VendorQuotation.query
    if f_status:
        query = query.filter(VendorQuotation.status == f_status)
    if f_vendor and str(f_vendor).isdigit():
        query = query.filter(VendorQuotation.vendor_id == int(f_vendor))

    total_count = query.count()
    total_pages = max(1, (total_count + per_page - 1) // per_page)
    page        = min(page, total_pages)
    quotes      = query.order_by(VendorQuotation.id.desc()).offset((page - 1) * per_page).limit(per_page).all()

    from flask import url_for as _uf
    base_url = _uf("admin_quotations", status=f_status, vendor_id=f_vendor)

    return render_template("admin_quotations.html", quotes=quotes,
        page=page, per_page=per_page,
        total_count=total_count, total_pages=total_pages,
        base_url=base_url,
        f_status=f_status, f_vendor=f_vendor)





# ==========================
# VENDOR BULK OPERATIONS
# ==========================
@app.route("/vendor-bulk", methods=["POST"])
@login_required
def vendor_bulk():
    """Legacy form-based bulk — kept for backward compat; redirects to vendor_list."""
    return redirect(url_for("vendor_list"))


@app.route("/vendor-bulk-ajax", methods=["POST"])
@login_required
def vendor_bulk_ajax():
    """AJAX bulk operations — returns JSON for enable/disable/delete; file for export."""
    import json as _json
    action  = request.form.get("action", "").strip()
    ids_raw = request.form.getlist("vendor_ids")
    if not ids_raw:
        return jsonify({"ok": False, "error": "No vendors selected."}), 400
    if not action:
        return jsonify({"ok": False, "error": "No action specified."}), 400
    ids     = [int(i) for i in ids_raw if str(i).isdigit()]
    vendors = Vendor.query.filter(Vendor.id.in_(ids)).all()
    count   = len(vendors)

    if action == "enable":
        for v in vendors:
            v.is_active = True
        db.session.commit()
        return jsonify({"ok": True, "message": f"Enabled {count} vendor(s) successfully.",
                        "action": "enable", "ids": ids})

    elif action == "disable":
        for v in vendors:
            v.is_active = False
        db.session.commit()
        return jsonify({"ok": True, "message": f"Disabled {count} vendor(s) successfully.",
                        "action": "disable", "ids": ids})

    elif action == "delete":
        for v in vendors:
            db.session.delete(v)
        db.session.commit()
        return jsonify({"ok": True, "message": f"Deleted {count} vendor(s) successfully.",
                        "action": "delete", "ids": ids})

    elif action == "export":
        wb = Workbook()
        ws = wb.active
        ws.title = "Vendors"
        headers = ["Company", "Owner", "Mobile", "Email", "District",
                   "Username", "Status", "Total Leads", "Completed Leads",
                   "Pending Leads", "Rating"]
        ws.append(headers)
        # Style header row
        from openpyxl.styles import Font, PatternFill, Alignment
        green_fill = PatternFill("solid", fgColor="005B38")
        for cell in ws[1]:
            cell.font      = Font(bold=True, color="FFFFFF")
            cell.fill      = green_fill
            cell.alignment = Alignment(horizontal="center")
        for v in vendors:
            total     = Lead.query.filter_by(vendor_id=v.id).count()
            completed = Lead.query.filter_by(vendor_id=v.id, status="Completed").count()
            pending   = Lead.query.filter(
                Lead.vendor_id == v.id,
                Lead.status.notin_(["Completed", "Cancelled"])
            ).count()
            ratings   = VendorRating.query.filter_by(vendor_id=v.id).all()
            avg_rating = round(sum(r.rating for r in ratings) / len(ratings), 1) if ratings else 0
            ws.append([
                v.company_name, v.owner_name, v.mobile, v.email or "",
                v.district, v.username,
                "Active" if v.is_active else "Inactive",
                total, completed, pending, avg_rating
            ])
        # Auto-width
        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return send_file(buf, as_attachment=True,
            download_name="vendors_export.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    return jsonify({"ok": False, "error": "Unknown action."}), 400


# ==========================
# VENDOR PRICING DATA API
# ==========================
@app.route("/vendor-pricing/get")
@vendor_required
def vendor_pricing_get():
    """Return pricing JSON for auto-fill in quotation builder."""
    vendor_id = session["vendor_id"]
    try:
        cap   = int(request.args.get("cap", 0))
        ptype = request.args.get("type", "Residential")
    except (ValueError, TypeError):
        return jsonify({"ok": False, "error": "Invalid parameters."}), 400

    p = VendorPricing.query.filter_by(
        vendor_id=vendor_id, project_type=ptype, capacity_kw=cap
    ).first()

    if not p:
        return jsonify({"ok": False, "error": "No pricing configured for this capacity and type."})

    # ── Component-wise sub-total ──────────────────────────────────────────
    panel_qty   = (p.panel_quantity or 0) if p.panel_quantity else max(1, cap)
    panel_total = round(float(p.panel_price or 0) * panel_qty, 2)
    inv_total   = float(p.inverter_price   or 0)
    str_total   = float(p.structure_price  or 0)
    comp_total  = round(
        panel_total + inv_total + str_total +
        float(p.install_charge   or 0) +
        float(p.elec_material    or 0) +
        float(p.net_meter_charge or 0) +
        float(p.transportation   or 0) +
        float(p.documentation    or 0) +
        float(p.miscellaneous    or 0),
        2
    )
    gst_pct = float(p.gst_percent or 0)
    gst_amt = round(comp_total * gst_pct / 100, 2)
    gross   = round(comp_total + gst_amt, 2)

    # ── Authoritative gross fallback: when component fields not filled,
    #    use vendor_price (what the vendor actually charges) ───────────────
    vendor_price = float(p.vendor_price or 0)
    if gross == 0 and vendor_price > 0:
        gst_amt = round(vendor_price * gst_pct / 100, 2)
        gross   = round(vendor_price + gst_amt, 2)
        comp_total = vendor_price   # treat vendor_price as the sub-total

    final_cust = round(vendor_price + float(p.hgs_commission or 0), 2)

    return jsonify({
        "ok":               True,
        "brand":            p.brand or "",
        "warranty_years":   int(p.warranty_years or 0),
        "vendor_price":     vendor_price,
        "hgs_commission":   float(p.hgs_commission or 0),
        "final_price":      float(p.final_price or 0) or final_cust,
        "panel_price":      float(p.panel_price or 0),
        "panel_wattage":    int(p.panel_wattage or 0),
        "panel_quantity":   panel_qty,
        "panel_total":      panel_total,
        "inverter_price":   inv_total,
        "inverter_brand":   p.inverter_brand or "",
        "inverter_capacity":p.inverter_capacity or "",
        "structure_price":  str_total,
        "structure_type":   p.structure_type or "",
        "install_charge":   float(p.install_charge   or 0),
        "elec_material":    float(p.elec_material    or 0),
        "net_meter_charge": float(p.net_meter_charge or 0),
        "transportation":   float(p.transportation   or 0),
        "documentation":    float(p.documentation    or 0),
        "miscellaneous":    float(p.miscellaneous    or 0),
        "gst_percent":      gst_pct,
        "comp_total":       comp_total,
        "gst_amount":       gst_amt,
        "gross_price":      gross,   # ← this is what JS sends back as form_gross
    })


# ==========================
# VENDOR CHANGE PASSWORD
# ==========================
@app.route("/vendor-change-password", methods=["POST"])
def vendor_change_password():
    if "vendor_id" not in session:
        return jsonify({"ok": False, "error": "Not logged in."}), 401
    vendor = Vendor.query.get_or_404(session["vendor_id"])
    current_pw  = request.form.get("current_password", "")
    new_pw      = request.form.get("new_password", "").strip()
    confirm_pw  = request.form.get("confirm_password", "").strip()
    if not check_password_hash(vendor.password, current_pw):
        return jsonify({"ok": False, "error": "Current password is incorrect."}), 400
    if len(new_pw) < 6:
        return jsonify({"ok": False, "error": "New password must be at least 6 characters."}), 400
    if new_pw != confirm_pw:
        return jsonify({"ok": False, "error": "New password and confirm password do not match."}), 400
    vendor.password = generate_password_hash(new_pw)
    db.session.commit()
    logger.info("Vendor %s changed password.", vendor.username)
    return jsonify({"ok": True, "message": "Password changed successfully."})


@app.route("/vendor-dashboard")
def vendor_dashboard():
    if "vendor_id" not in session:
        return redirect("/vendor-login")

    vendor_id = session["vendor_id"]
    vendor    = Vendor.query.get_or_404(vendor_id)

    f_status = request.args.get("status", "")
    f_search = request.args.get("search", "")

    query = Lead.query.filter_by(vendor_id=vendor_id)
    if f_status:
        query = query.filter(Lead.status == f_status)
    if f_search:
        like = f"%{f_search}%"
        query = query.filter(db.or_(
            Lead.name.like(like), Lead.phone.like(like)
        ))

    leads = query.order_by(Lead.id.desc()).all()

    total_leads     = Lead.query.filter_by(vendor_id=vendor_id).count()
    completed_leads = Lead.query.filter_by(vendor_id=vendor_id, status="Completed").count()
    pending_leads   = Lead.query.filter(
        Lead.vendor_id == vendor_id,
        Lead.status.notin_(["Completed", "Cancelled"])
    ).count()
    cancelled_leads = Lead.query.filter_by(vendor_id=vendor_id, status="Cancelled").count()
    completion_pct  = round((completed_leads / total_leads * 100), 1) if total_leads else 0

    today = datetime.now().strftime("%Y-%m-%d")
    today_followups = Lead.query.filter_by(
        vendor_id=vendor_id, follow_date=today
    ).all()

    # new feature counts
    pending_quotations = VendorQuotation.query.filter_by(
        vendor_id=vendor_id, status="Pending").count()

    return render_template(
        "vendor_dashboard.html",
        vendor=vendor,
        leads=leads,
        STATUSES=LEAD_STATUSES,
        total_leads=total_leads,
        completed_leads=completed_leads,
        pending_leads=pending_leads,
        cancelled_leads=cancelled_leads,
        completion_pct=completion_pct,
        today_followups=today_followups,
        pending_quotations=pending_quotations,
        f_status=f_status,
        f_search=f_search
    )

# ==========================
# VENDOR LIST (admin)
# ==========================
# VENDOR LIST (admin)
# ==========================
@app.route("/vendors")
@login_required
def vendor_list():
    f_district = request.args.get("district", "")
    f_active   = request.args.get("active", "")
    try:
        page     = max(1, int(request.args.get("page", 1)))
        per_page = int(request.args.get("per_page", 25))
        if per_page not in (10, 25, 50, 100):
            per_page = 25
    except (ValueError, TypeError):
        page = 1; per_page = 25

    query = Vendor.query
    if f_district:
        query = query.filter(Vendor.district == f_district)
    if f_active != "":
        query = query.filter(Vendor.is_active == (f_active == "1"))

    total_count = query.count()
    total_pages = max(1, (total_count + per_page - 1) // per_page)
    page        = min(page, total_pages)
    vendors     = query.order_by(Vendor.id.desc()).offset((page - 1) * per_page).limit(per_page).all()

    perf = {}
    for v in vendors:
        total     = Lead.query.filter_by(vendor_id=v.id).count()
        completed = Lead.query.filter_by(vendor_id=v.id, status="Completed").count()
        pending   = Lead.query.filter(
            Lead.vendor_id == v.id,
            Lead.status.notin_(["Completed", "Cancelled"])
        ).count()
        cancelled = Lead.query.filter_by(vendor_id=v.id, status="Cancelled").count()
        pct       = round((completed / total * 100), 1) if total else 0
        perf[v.id] = {
            "total": total, "completed": completed,
            "pending": pending, "cancelled": cancelled, "pct": pct
        }

    # Build base_url for pagination links (preserve filters, strip page/per_page)
    from flask import url_for as _uf
    base_url = _uf("vendor_list", district=f_district, active=f_active)

    return render_template(
        "vendor_list.html",
        vendors=vendors, perf=perf,
        districts=UP_DISTRICTS,
        f_district=f_district, f_active=f_active,
        page=page, per_page=per_page,
        total_count=total_count, total_pages=total_pages,
        base_url=base_url
    )

# ==========================
# VENDOR DETAIL (admin)
# ==========================
@app.route("/vendor-detail/<int:vendor_id>")
@login_required
def vendor_detail(vendor_id):
    import json as _json, statistics as _stat
    v        = Vendor.query.get_or_404(vendor_id)
    profile  = VendorProfile.query.filter_by(vendor_id=vendor_id).first()
    # Guard: if vendor has no profile yet, provide an in-memory default
    # so the template never crashes on profile.field access
    if not profile:
        profile = VendorProfile(vendor_id=vendor_id)
    pricing  = VendorPricing.query.filter_by(vendor_id=vendor_id).order_by(
                    VendorPricing.capacity_kw, VendorPricing.project_type).all()
    commissions = VendorCommission.query.filter_by(vendor_id=vendor_id).order_by(
                    VendorCommission.capacity_kw).all()
    coverage = VendorCoverage.query.filter_by(vendor_id=vendor_id).first()
    quotations = VendorQuotation.query.filter_by(vendor_id=vendor_id)\
                    .order_by(VendorQuotation.id.desc()).all()
    ratings  = VendorRating.query.filter_by(vendor_id=vendor_id).all()

    # Lead performance
    all_leads   = Lead.query.filter_by(vendor_id=vendor_id).all()
    total_l     = len(all_leads)
    completed_l = sum(1 for l in all_leads if l.status == "Completed")
    pending_l   = sum(1 for l in all_leads if l.status not in ("Completed","Cancelled"))
    cancelled_l = sum(1 for l in all_leads if l.status == "Cancelled")
    conv_pct    = round(completed_l / total_l * 100, 1) if total_l else 0

    # Timeline: last 15 timeline entries for this vendor's leads
    lead_ids = [l.id for l in all_leads]
    timeline = []
    if lead_ids:
        timeline = LeadTimeline.query.filter(LeadTimeline.lead_id.in_(lead_ids))\
                       .order_by(LeadTimeline.id.desc()).limit(15).all()

    # Ratings
    avg_rating = round(_stat.mean(r.rating for r in ratings), 1) if ratings else 0

    # Revenue from approved quotations
    approved_q  = [q for q in quotations if q.status == "Approved"]
    total_rev   = sum(q.net_price for q in approved_q)
    total_comm  = sum(q.commission for q in approved_q)

    # Coverage parsing
    cov_districts = []
    cov_types     = []
    if coverage:
        try: cov_districts = _json.loads(coverage.districts or "[]")
        except: cov_districts = []
        try: cov_types = _json.loads(coverage.lead_types or "[]")
        except: cov_types = []

    # Brands from profile
    brands = []
    if profile and profile.brands_supported:
        brands = [b.strip() for b in profile.brands_supported.split(",") if b.strip()]

    # Status breakdown
    status_counts = {}
    for l in all_leads:
        status_counts[l.status] = status_counts.get(l.status, 0) + 1

    return render_template("vendor_detail.html",
        v=v, profile=profile, pricing=pricing,
        commissions=commissions, coverage=coverage,
        quotations=quotations, ratings=ratings,
        timeline=timeline,
        total_l=total_l, completed_l=completed_l,
        pending_l=pending_l, cancelled_l=cancelled_l,
        conv_pct=conv_pct, avg_rating=avg_rating,
        total_rev=total_rev, total_comm=total_comm,
        cov_districts=cov_districts, cov_types=cov_types,
        brands=brands, status_counts=status_counts,
        capacities=KW_CAPACITIES,
    )


# ==========================
# VENDOR COMPARISON (admin)
# ==========================
@app.route("/vendor-compare")
@login_required
def vendor_compare():
    import statistics as _stat
    ids_raw = request.args.getlist("ids")
    district_filter = request.args.get("district", "")

    # If single vendor id + district requested, auto-load all peers from that district
    if district_filter and len(ids_raw) == 1:
        anchor_id = int(ids_raw[0])
        peers = Vendor.query.filter_by(district=district_filter, is_active=True)\
                            .filter(Vendor.is_deleted != True).all()
        ids_raw = [str(anchor_id)] + [str(p.id) for p in peers if p.id != anchor_id]

    if not ids_raw:
        defaults = Vendor.query.filter(Vendor.is_deleted != True)\
                               .order_by(Vendor.id).limit(3).all()
        ids_raw = [str(v.id) for v in defaults]

    ids = [int(i) for i in ids_raw if str(i).isdigit()][:6]

    # BUG 7 FIX: eager-load relationships to avoid N+1 queries
    from sqlalchemy.orm import joinedload
    vendors = Vendor.query.options(
        joinedload(Vendor.profile),
        joinedload(Vendor.pricing),
        joinedload(Vendor.commissions),
        joinedload(Vendor.ratings),
    ).filter(Vendor.id.in_(ids)).all()

    # Pagination for vendor selector
    try:
        sel_page     = max(1, int(request.args.get("sel_page", 1)))
        sel_per_page = 20
    except (ValueError, TypeError):
        sel_page = 1; sel_per_page = 20

    all_vendors_q     = Vendor.query.filter_by(is_active=True)\
                                    .filter(Vendor.is_deleted != True)\
                                    .order_by(Vendor.company_name)
    all_vendors_total = all_vendors_q.count()
    all_vendors_pages = max(1, (all_vendors_total + sel_per_page - 1) // sel_per_page)
    sel_page          = min(sel_page, all_vendors_pages)
    all_vendors       = all_vendors_q.offset((sel_page - 1) * sel_per_page)\
                                     .limit(sel_per_page).all()

    # Pre-fetch all lead counts in one query to avoid N+1
    lead_counts = db.session.query(
        Lead.vendor_id,
        db.func.count(Lead.id).label("total"),
        db.func.sum(db.case((Lead.status == "Completed", 1), else_=0)).label("completed")
    ).filter(Lead.vendor_id.in_(ids)).group_by(Lead.vendor_id).all()
    lead_map = {r.vendor_id: (r.total or 0, r.completed or 0) for r in lead_counts}

    compare_data = []
    for v in vendors:
        total_l, completed = lead_map.get(v.id, (0, 0))
        conv_pct = round(completed / total_l * 100, 1) if total_l else 0

        # Ratings from eager-loaded relationship
        ratings = v.ratings or []
        avg_r   = round(_stat.mean(r.rating for r in ratings), 1) if ratings else 0

        # Approved quotations — fetch directly
        approved  = VendorQuotation.query.filter_by(vendor_id=v.id, status="Approved").all()
        total_rev = sum(q.net_price  for q in approved)
        total_com = sum(q.commission for q in approved)

        # Pricing maps from eager-loaded relationship
        pricing_res     = [p for p in (v.pricing or []) if p.project_type == "Residential"]
        pricing_com     = [p for p in (v.pricing or []) if p.project_type == "Commercial"]
        pricing_gov     = [p for p in (v.pricing or []) if p.project_type == "Government"]
        pricing_map_res = {p.capacity_kw: p for p in pricing_res}
        pricing_map_com = {p.capacity_kw: p for p in pricing_com}
        pricing_map_gov = {p.capacity_kw: p for p in pricing_gov}

        # Average commission from pricing (residential) — use eager-loaded data
        res_with_price = [p for p in pricing_res if p.hgs_commission]
        avg_comm = round(sum(p.hgs_commission for p in res_with_price) / len(res_with_price), 0) \
                   if res_with_price else 0

        # Representative residential price (1 KW or first available)
        res_price = (pricing_map_res.get(1) or (pricing_res[0] if pricing_res else None))
        com_price = (pricing_map_com.get(1) or (pricing_com[0] if pricing_com else None))
        gov_price = (pricing_map_gov.get(1) or (pricing_gov[0] if pricing_gov else None))

        # Profile from eager-loaded relationship
        profile = v.profile
        brands  = []
        if profile and profile.brands_supported:
            brands = [b.strip() for b in profile.brands_supported.split(",") if b.strip()]

        warranty     = pricing_res[0].warranty_years if pricing_res else 0
        updated_date = (profile.updated_at if profile and profile.updated_at else "—")

        compare_data.append({
            "vendor":          v,
            "profile":         profile,
            "total_leads":     total_l,
            "completed":       completed,
            "pending":         total_l - completed,
            "conv_pct":        conv_pct,
            "avg_rating":      avg_r,
            "total_rev":       total_rev,
            "total_comm":      total_com,
            "avg_comm":        avg_comm,
            "pricing_map":     pricing_map_res,
            "pricing_map_com": pricing_map_com,
            "pricing_map_gov": pricing_map_gov,
            "res_price":       res_price,
            "com_price":       com_price,
            "gov_price":       gov_price,
            "brands":          brands,
            "warranty":        warranty,
            "updated_date":    updated_date,
        })

    # Determine best vendor by composite score
    if compare_data:
        for row in compare_data:
            row["score"] = row["conv_pct"] + row["avg_rating"] * 10
        best_id = max(compare_data, key=lambda x: x["score"])["vendor"].id
    else:
        best_id = None

    return render_template("vendor_compare.html",
        compare_data=compare_data,
        all_vendors=all_vendors,
        selected_ids=ids,
        capacities=KW_CAPACITIES,
        best_id=best_id,
        district_filter=district_filter,
        sel_page=sel_page,
        sel_per_page=sel_per_page,
        all_vendors_total=all_vendors_total,
        all_vendors_pages=all_vendors_pages,
    )


# ==========================
# VENDOR ANALYTICS (admin)
# ==========================
@app.route("/vendor-analytics")
@login_required
def vendor_analytics():
    import statistics as _stat
    from sqlalchemy.orm import joinedload

    # BUG 7 FIX: eager-load to avoid N+1 queries
    vendors = Vendor.query.options(
        joinedload(Vendor.profile),
        joinedload(Vendor.pricing),
        joinedload(Vendor.ratings),
    ).all()

    # ── Summary stats (BUG 4 FIX: real DB values, no hardcodes) ──
    total_vendors    = len(vendors)
    active_vendors   = sum(1 for v in vendors if v.is_active and not v.is_deleted)
    disabled_vendors = sum(1 for v in vendors if not v.is_active and not v.is_deleted)

    # PM Surya / DISCOM counts from VendorProfile
    pm_surya_count  = VendorProfile.query.filter_by(pm_surya_approved=True).count()
    discom_count    = VendorProfile.query.filter_by(discom_empanelled=True).count()

    # Vendors with residential / commercial / government pricing
    res_vendor_ids  = db.session.query(VendorPricing.vendor_id.distinct())\
                                .filter_by(project_type="Residential").count()
    com_vendor_ids  = db.session.query(VendorPricing.vendor_id.distinct())\
                                .filter_by(project_type="Commercial").count()
    gov_vendor_ids  = db.session.query(VendorPricing.vendor_id.distinct())\
                                .filter_by(project_type="Government").count()

    # Average commission and price from residential pricing
    all_res_pricing = VendorPricing.query.filter(
        VendorPricing.project_type == "Residential",
        VendorPricing.final_price > 0
    ).all()
    avg_commission = round(
        sum(p.hgs_commission for p in all_res_pricing) / len(all_res_pricing), 0
    ) if all_res_pricing else 0
    avg_price = round(
        sum(p.final_price for p in all_res_pricing) / len(all_res_pricing), 0
    ) if all_res_pricing else 0

    # Top districts by vendor count
    _dist_rows = db.session.query(
        Vendor.district, db.func.count(Vendor.id).label("cnt")
    ).filter(Vendor.is_deleted != True, Vendor.district != "")\
     .group_by(Vendor.district).order_by(db.func.count(Vendor.id).desc()).limit(10).all()
    top_districts = [{"district": r[0], "count": r[1]} for r in _dist_rows]

    # Vendor registration trend (last 12 months by created_at)
    now = datetime.now()
    reg_trend_labels, reg_trend_counts = [], []
    for i in range(11, -1, -1):
        d   = now - timedelta(days=i * 30)
        ym  = d.strftime("%Y-%m")
        lbl = d.strftime("%b %Y")
        cnt = sum(1 for v in vendors if v.created_at and v.created_at[:7] == ym)
        reg_trend_labels.append(lbl)
        reg_trend_counts.append(cnt)

    # Pre-fetch all lead counts in a single aggregated query
    lead_counts_q = db.session.query(
        Lead.vendor_id,
        db.func.count(Lead.id).label("total"),
        db.func.sum(db.case((Lead.status == "Completed", 1), else_=0)).label("completed")
    ).group_by(Lead.vendor_id).all()
    lead_map = {r.vendor_id: (r.total or 0, r.completed or 0) for r in lead_counts_q}

    rows = []
    for v in vendors:
        total_l, completed = lead_map.get(v.id, (0, 0))
        conv_pct = round(completed / total_l * 100, 1) if total_l else 0

        # Use eager-loaded ratings
        ratings = v.ratings or []
        avg_r   = round(_stat.mean(r.rating for r in ratings), 1) if ratings else 0

        # Revenue from approved quotations (separate query but batched)
        approved  = VendorQuotation.query.filter_by(vendor_id=v.id, status="Approved").all()
        total_rev = sum(q.net_price  for q in approved)
        total_com = sum(q.commission for q in approved)

        # Use eager-loaded pricing
        res_prices = [p for p in (v.pricing or [])
                      if p.project_type == "Residential" and p.final_price > 0]
        min_price = min((p.final_price for p in res_prices), default=0)
        avg_comm  = round(
            sum(p.hgs_commission for p in res_prices) / len(res_prices), 0
        ) if res_prices else 0

        rows.append({
            "vendor":     v,
            "total_leads": total_l,
            "completed":  completed,
            "conv_pct":   conv_pct,
            "avg_rating": avg_r,
            "total_rev":  total_rev,
            "total_comm": total_com,
            "min_price":  min_price,
            "avg_comm":   avg_comm,
        })

    top_commission = sorted(rows, key=lambda x: -x["avg_comm"])[:10]
    top_low_price  = [r for r in sorted(rows, key=lambda x: x["min_price"])
                      if r["min_price"] > 0][:10]
    top_conversion = sorted(rows, key=lambda x: -x["conv_pct"])[:10]
    top_rated      = sorted(rows, key=lambda x: -x["avg_rating"])[:10]
    top_revenue    = sorted(rows, key=lambda x: -x["total_rev"])[:10]
    top_completed  = sorted(rows, key=lambda x: -x["completed"])[:10]
    most_active    = sorted(rows, key=lambda x: -x["total_leads"])[:10]

    return render_template("vendor_analytics.html",
        # Summary stats
        total_vendors=total_vendors,
        active_vendors=active_vendors,
        disabled_vendors=disabled_vendors,
        pm_surya_count=pm_surya_count,
        discom_count=discom_count,
        res_vendor_count=res_vendor_ids,
        com_vendor_count=com_vendor_ids,
        gov_vendor_count=gov_vendor_ids,
        avg_commission=avg_commission,
        avg_price=avg_price,
        top_districts=top_districts,
        reg_trend_labels=reg_trend_labels,
        reg_trend_counts=reg_trend_counts,
        # Rankings
        top_commission=top_commission,
        top_low_price=top_low_price,
        top_conversion=top_conversion,
        top_rated=top_rated,
        top_revenue=top_revenue,
        top_completed=top_completed,
        most_active=most_active,
    )


# ==========================
# ADD VENDOR (admin)
# ==========================
@app.route("/vendor-add", methods=["GET", "POST"])
@login_required
def vendor_add():
    if request.method == "POST":
        company_name = request.form.get("company_name", "").strip()
        owner_name   = request.form.get("owner_name", "").strip()
        mobile       = request.form.get("mobile", "").strip()
        email        = request.form.get("email", "").strip()
        district     = request.form.get("district", "").strip()

        # Auto-generate username & password if not provided
        base     = owner_name.lower().replace(" ", ".")[:12] or "vendor"
        username = request.form.get("username", "").strip()
        if not username:
            username = f"{base}.{secrets.token_hex(3)}"
            while Vendor.query.filter_by(username=username).first():
                username = f"{base}.{secrets.token_hex(3)}"

        raw_pw = request.form.get("password", "").strip()
        if not raw_pw:
            raw_pw = secrets.token_urlsafe(10)

        if Vendor.query.filter_by(username=username).first():
            flash("Username already exists.", "danger")
            return render_template("vendor_add.html", districts=UP_DISTRICTS)

        vendor = Vendor(
            company_name=company_name,
            owner_name=owner_name,
            mobile=mobile,
            email=email,
            username=username,
            password=generate_password_hash(raw_pw),
            district=district,
            is_active=True,
            created_at=datetime.now().strftime("%d-%m-%Y")
        )
        db.session.add(vendor)
        db.session.commit()

        # Check if this is an AJAX/JSON request (from credential modal)
        if request.headers.get("X-Requested-With") == "XMLHttpRequest":
            import json
            return json.dumps({
                "ok":           True,
                "company_name": company_name,
                "owner_name":   owner_name,
                "username":     username,
                "password":     raw_pw,
                "district":     district,
                "email":        email,
                "mobile":       mobile
            }), 200, {"Content-Type": "application/json"}

        flash(f"Vendor '{company_name}' added successfully.", "success")
        return redirect(url_for("vendor_list"))
    return render_template("vendor_add.html", districts=UP_DISTRICTS)

# ==========================
# EDIT VENDOR (admin)
# ==========================
@app.route("/vendor-edit/<int:vendor_id>", methods=["GET", "POST"])
@admin_required
def vendor_edit(vendor_id):
    vendor = Vendor.query.filter(Vendor.id == vendor_id, Vendor.is_deleted != True)\
                         .first_or_404()
    if request.method == "POST":
        vendor.is_active     = request.form.get("is_active") == "1"
        vendor.approved_date = request.form.get("approved_date", vendor.approved_date).strip()
        vendor.district      = request.form.get("district", vendor.district)
        vendor.vendor_code   = request.form.get("vendor_code", vendor.vendor_code or "").strip()
        vendor.admin_remarks = request.form.get("admin_remarks", "").strip()
        db.session.commit()
        flash("Vendor updated successfully.", "success")
        return redirect(url_for("vendor_detail", vendor_id=vendor_id))
    return render_template("vendor_edit.html", vendor=vendor, districts=UP_DISTRICTS)

# ==========================
# DELETE VENDOR (admin)
# ==========================
@app.route("/vendor-delete/<int:vendor_id>")
@admin_required
def vendor_delete(vendor_id):
    vendor = Vendor.query.get_or_404(vendor_id)
    Lead.query.filter_by(vendor_id=vendor_id).update(
        {"vendor_id": None, "status": "New"}
    )
    db.session.delete(vendor)
    db.session.commit()
    flash("Vendor deleted.", "success")
    return redirect(url_for("vendor_list"))

# ==========================
# TOGGLE VENDOR STATUS
# ==========================
@app.route("/vendor-toggle/<int:vendor_id>")
@admin_required
def vendor_toggle(vendor_id):
    vendor = Vendor.query.get_or_404(vendor_id)
    vendor.is_active = not vendor.is_active
    db.session.commit()
    flash(f"Vendor {'enabled' if vendor.is_active else 'disabled'}.", "success")
    return redirect(url_for("vendor_list"))

# ==========================
# SUPER-ADMIN: RESET VENDOR PASSWORD
# ==========================
@app.route("/vendor-reset-password/<int:vendor_id>", methods=["POST"])
@login_required
def vendor_reset_password(vendor_id):
    """Admin only: auto-generate a strong temporary password, hash + save it, return plain text once."""
    import secrets, string as _string
    user = User.query.get(session.get("user_id"))
    if not user or session.get("role") != "admin":
        return jsonify({"ok": False, "error": "Admin access required."}), 403
    vendor = Vendor.query.filter(Vendor.id == vendor_id, Vendor.is_deleted != True).first_or_404()
    # Generate strong 12-char temp password
    alphabet  = _string.ascii_letters + _string.digits + "!@#$%"
    temp_pw   = "".join(secrets.choice(alphabet) for _ in range(12))
    vendor.password = generate_password_hash(temp_pw)
    db.session.commit()
    return jsonify({
        "ok": True,
        "temp_password": temp_pw,
        "username": vendor.username,
        "company": vendor.company_name,
        "email": vendor.email or "",
        "mobile": vendor.mobile,
        "message": f"Password for {vendor.company_name} has been reset."
    })

# ==========================
# SUPER-ADMIN: ENABLE / DISABLE VENDOR (JSON)
# ==========================
@app.route("/vendor-setstatus/<int:vendor_id>", methods=["POST"])
@login_required
def vendor_setstatus(vendor_id):
    """Admin: enable or disable a vendor via JSON endpoint."""
    if session.get("role") != "admin":
        return jsonify({"ok": False, "error": "Admin access required."}), 403
    vendor = Vendor.query.filter(Vendor.id == vendor_id, Vendor.is_deleted != True).first_or_404()
    action = request.form.get("action", "")
    if action == "enable":
        vendor.is_active = True
    elif action == "disable":
        vendor.is_active = False
    else:
        return jsonify({"ok": False, "error": "Invalid action."}), 400
    db.session.commit()
    return jsonify({"ok": True, "is_active": vendor.is_active,
                    "message": f"Vendor {'enabled' if vendor.is_active else 'disabled'}."})

# ==========================
# SUPER-ADMIN: DELETE VENDOR (JSON)
# ==========================
@app.route("/vendor-delete-ajax/<int:vendor_id>", methods=["POST"])
@login_required
def vendor_delete_ajax(vendor_id):
    """Admin: soft-delete vendor (keeps quotations + lead history)."""
    if session.get("role") != "admin":
        return jsonify({"ok": False, "error": "Admin access required."}), 403
    vendor = Vendor.query.get_or_404(vendor_id)
    vendor.is_deleted = True
    vendor.is_active  = False
    # Unassign pending leads but keep history
    Lead.query.filter(Lead.vendor_id == vendor_id,
                      Lead.status.notin_(["Completed", "Cancelled"]))\
              .update({"vendor_id": None, "status": "New"}, synchronize_session=False)
    db.session.commit()
    return jsonify({"ok": True, "message": f"Vendor {vendor.company_name} has been deleted."})

# ==========================
# IMPORT VENDORS FROM EXCEL
# ==========================
@app.route("/import-vendors", methods=["GET", "POST"])
@admin_required
def import_vendors():
    if request.method == "POST":
        file = request.files.get("file")
        if not file:
            flash("No file selected.", "danger")
            return redirect("/import-vendors")
        try:
            df = pd.read_excel(file)

            # Normalize column names: strip whitespace + lowercase
            df.columns = [str(c).strip() for c in df.columns]

            # Column alias map — supports both old template and UP approved vendor list
            COL_ALIASES = {
                "company_name":  ["company_name", "firm name", "firm_name"],
                "mobile":        ["mobile", "mobile number", "mobile_number", "phone"],
                "district":      ["district"],
                "email":         ["email", "email id", "email_id"],
                "owner_name":    ["owner_name", "owner name"],
                "vendor_code":   ["vendor code", "vendor_code"],
                "address":       ["address"],
                "approved_date": ["approved date", "approved_date"],
            }

            def find_col(df_cols, aliases):
                df_lower = {c.lower(): c for c in df_cols}
                for alias in aliases:
                    if alias.lower() in df_lower:
                        return df_lower[alias.lower()]
                return None

            col_map = {field: find_col(df.columns, aliases) for field, aliases in COL_ALIASES.items()}

            # company_name and mobile are mandatory
            if not col_map["company_name"]:
                flash("Could not find Firm Name / company_name column in the Excel file.", "danger")
                return redirect("/import-vendors")
            if not col_map["mobile"]:
                flash("Could not find Mobile Number / mobile column in the Excel file.", "danger")
                return redirect("/import-vendors")

            def get_val(row, field):
                c = col_map.get(field)
                if not c or c not in row or pd.isna(row[c]):
                    return ""
                return str(row[c]).strip()

            added = 0
            skipped = 0
            failed = 0

            for _, row in df.iterrows():
                try:
                    company  = get_val(row, "company_name")
                    mobile   = get_val(row, "mobile")
                    # Clean mobile: keep digits only, take last 10
                    mobile_clean = "".join(filter(str.isdigit, mobile))
                    if len(mobile_clean) >= 10:
                        mobile_clean = mobile_clean[-10:]

                    if not company or not mobile_clean:
                        skipped += 1
                        continue

                    # Skip duplicate mobile numbers
                    if Vendor.query.filter_by(mobile=mobile_clean).first():
                        skipped += 1
                        continue

                    district      = get_val(row, "district")
                    email         = get_val(row, "email")
                    owner         = get_val(row, "owner_name") or company  # fallback to firm name
                    vendor_code   = get_val(row, "vendor_code")
                    address       = get_val(row, "address")
                    approved_date = get_val(row, "approved_date")

                    # Auto-generate unique username
                    username = f"vendor_{mobile_clean}"
                    if Vendor.query.filter_by(username=username).first():
                        skipped += 1
                        continue

                    vendor = Vendor(
                        company_name=company,
                        owner_name=owner,
                        mobile=mobile_clean,
                        email=email,
                        username=username,
                        password=generate_password_hash(mobile_clean),
                        district=district,
                        vendor_code=vendor_code,
                        address=address,
                        approved_date=approved_date,
                        is_active=True,
                        created_at=datetime.now().strftime("%d-%m-%Y")
                    )
                    db.session.add(vendor)
                    added += 1
                except Exception:
                    failed += 1
                    continue

            db.session.commit()
            parts = [f"{added} imported", f"{skipped} skipped"]
            if failed:
                parts.append(f"{failed} failed")
            flash(f"Import complete: {', '.join(parts)}.", "success")
        except Exception as e:
            flash(f"Import failed: {str(e)}", "danger")
        return redirect("/vendors")
    return render_template("import_vendors.html")

# ==========================
# PROFESSIONAL EXCEL EXPORT
# ==========================
@app.route("/download-leads")
@login_required
def download():

    leads = Lead.query.order_by(Lead.id.desc()).all()

    data = []
    for lead in leads:
        data.append({
            "ID": lead.id,
            "Customer Name": lead.name,
            "Mobile Number": lead.phone,
            "City": lead.city,
            "District": lead.district,
            "Monthly Bill": lead.bill,
            "Status": lead.status,
            "Assigned Vendor": lead.vendor.company_name if lead.vendor else "",
            "Follow Note": lead.note,
            "Follow Date": lead.follow_date,
            "Updated By": lead.updated_by,
            "Created Date": lead.created_at
        })

    filename = "Har_Ghar_Solar_CRM.xlsx"
    df = pd.DataFrame(data)
    df.to_excel(filename, index=False, startrow=5)

    wb = load_workbook(filename)
    ws = wb.active
    ws.title = "Solar Leads"

    # Branding header
    num_cols = len(data[0]) if data else 12
    last_col = get_column_letter(num_cols)

    ws.merge_cells(f"A1:{last_col}1")
    ws["A1"] = "HAR GHAR SOLAR — CRM Report"
    ws["A1"].font      = Font(bold=True, size=20, color="FFFFFF")
    ws["A1"].fill      = PatternFill("solid", fgColor="005B38")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35

    ws.merge_cells(f"A2:{last_col}2")
    ws["A2"] = f"Generated: {datetime.now().strftime('%d-%m-%Y %H:%M')}  |  Total Leads: {len(leads)}"
    ws["A2"].font      = Font(size=11, color="FFFFFF")
    ws["A2"].fill      = PatternFill("solid", fgColor="007A3D")
    ws["A2"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 22

    # Status summary
    ws.merge_cells(f"A3:{last_col}3")
    summary = "  |  ".join([f"{s}: {Lead.query.filter_by(status=s).count()}" for s in LEAD_STATUSES])
    ws["A3"] = summary
    ws["A3"].font      = Font(size=10, bold=True, color="005B38")
    ws["A3"].fill      = PatternFill("solid", fgColor="E8F5E9")
    ws["A3"].alignment = Alignment(horizontal="center")

    ws.merge_cells(f"A4:{last_col}4")
    ws["A4"] = ""

    # Header row style
    status_col_idx = None
    for i, cell in enumerate(ws[6], 1):
        cell.font      = Font(bold=True, color="FFFFFF", size=11)
        cell.fill      = PatternFill("solid", fgColor="1B5E20")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if cell.value == "Status":
            status_col_idx = i
    ws.row_dimensions[6].height = 24

    # Status color map
    status_colors = {
        "New":           "FFF9C4",
        "Assigned":      "E3F2FD",
        "Contacted":     "E8EAF6",
        "Site Visit":    "F3E5F5",
        "Quotation Sent":"FFF3E0",
        "Installation":  "E0F2F1",
        "Completed":     "C8E6C9",
        "Cancelled":     "FFCDD2"
    }

    border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    for row in ws.iter_rows(min_row=6):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if cell.row % 2 == 0 and cell.row > 6:
                if not (status_col_idx and cell.column == status_col_idx):
                    cell.fill = PatternFill("solid", fgColor="F1F8F4")

    # Status badge coloring
    if status_col_idx:
        for row in ws.iter_rows(min_row=7, min_col=status_col_idx, max_col=status_col_idx):
            for cell in row:
                val = str(cell.value) if cell.value else ""
                if val in status_colors:
                    cell.fill = PatternFill("solid", fgColor=status_colors[val])
                    cell.font = Font(bold=True, size=10)

    # Column widths
    col_widths = {
        "A": 6, "B": 22, "C": 15, "D": 15, "E": 18,
        "F": 18, "G": 16, "H": 24, "I": 28, "J": 14,
        "K": 16, "L": 14
    }
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width
    ws.row_dimensions[6].height = 24
    for r in range(7, ws.max_row + 1):
        ws.row_dimensions[r].height = 18

    # Excel table
    try:
        last_row    = ws.max_row
        table_range = f"A6:{last_col}{last_row}"
        tbl = Table(displayName="SolarLeads", ref=table_range)
        tbl.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium7",
            showFirstColumn=False, showLastColumn=False,
            showRowStripes=True, showColumnStripes=False
        )
        ws.add_table(tbl)
    except Exception as e:
        print("TABLE ERROR:", e)

    ws.freeze_panes = "A7"
    wb.save(filename)

    return send_file(filename, as_attachment=True, download_name=filename)


# ==========================
# LEAD DETAIL API (JSON)
# ==========================
@app.route("/lead-detail/<int:lead_id>")
@login_required
def lead_detail(lead_id):
    lead = Lead.query.get_or_404(lead_id)
    return jsonify({
        "id":          lead.id,
        "name":        lead.name or "",
        "phone":       lead.phone or "",
        "city":        lead.city or "",
        "district":    lead.district or "",
        "bill":        lead.bill or "",
        "status":      lead.status or "New",
        "note":        lead.note or "",
        "follow_date": lead.follow_date or "",
        "updated_by":  lead.updated_by or "",
        "created_at":  lead.created_at or "",
        "vendor_id":   lead.vendor_id or "",
        "vendor_name": lead.vendor.company_name if lead.vendor else ""
    })

# ==========================
# LEAD EDIT (POST — saves edits)
# ==========================
@app.route("/lead-edit/<int:lead_id>", methods=["POST"])
@login_required
def lead_edit(lead_id):
    lead = Lead.query.get_or_404(lead_id)
    old_status = lead.status
    lead.name        = request.form.get("name",        lead.name)
    lead.phone       = request.form.get("phone",       lead.phone)
    lead.city        = request.form.get("city",        lead.city)
    lead.district    = request.form.get("district",    lead.district)
    lead.bill        = request.form.get("bill",        lead.bill)
    lead.status      = request.form.get("status",      lead.status)
    lead.note        = request.form.get("note",        lead.note)
    lead.follow_date = request.form.get("follow_date", lead.follow_date)
    new_vendor_id    = request.form.get("vendor_id",   "")
    lead.vendor_id   = int(new_vendor_id) if new_vendor_id else None
    actor            = session.get("username", "Admin")
    lead.updated_by  = actor
    db.session.commit()
    if lead.status != old_status:
        add_timeline(lead.id, f"Status Changed: {old_status} → {lead.status}", "", actor)
    add_timeline(lead.id, "Lead Updated via Edit", "", actor)
    return jsonify({"success": True, "message": "Lead updated successfully."})

# ============================================================
# SUPER ADMIN — ADMIN MANAGEMENT
# ============================================================

def _generate_strong_password():
    """Generate a strong temporary password like HGS@8392XZ."""
    import random, string
    prefix    = "HGS@"
    digits    = ''.join(random.choices(string.digits, k=4))
    letters   = ''.join(random.choices(string.ascii_uppercase, k=2))
    return prefix + digits + letters

@app.route("/admin-management")
@super_admin_required
def admin_management():
    # ── Users pagination ──
    try:
        u_page     = max(1, int(request.args.get("u_page", 1)))
        u_per_page = int(request.args.get("u_per_page", 25))
        if u_per_page not in (10, 25, 50, 100):
            u_per_page = 25
    except (ValueError, TypeError):
        u_page = 1; u_per_page = 25

    u_total       = User.query.count()
    u_total_pages = max(1, (u_total + u_per_page - 1) // u_per_page)
    u_page        = min(u_page, u_total_pages)
    users         = User.query.order_by(User.id.asc())\
                        .offset((u_page - 1) * u_per_page).limit(u_per_page).all()

    # ── Vendors pagination ──
    try:
        v_page     = max(1, int(request.args.get("v_page", 1)))
        v_per_page = int(request.args.get("v_per_page", 25))
        if v_per_page not in (10, 25, 50, 100):
            v_per_page = 25
    except (ValueError, TypeError):
        v_page = 1; v_per_page = 25

    v_total       = Vendor.query.count()
    v_total_pages = max(1, (v_total + v_per_page - 1) // v_per_page)
    v_page        = min(v_page, v_total_pages)
    vendors       = Vendor.query.order_by(Vendor.id.asc())\
                        .offset((v_page - 1) * v_per_page).limit(v_per_page).all()
    for v in vendors:
        v._total_leads     = Lead.query.filter_by(vendor_id=v.id).count()
        v._completed_leads = Lead.query.filter_by(vendor_id=v.id, status="Completed").count()

    from flask import url_for as _uf
    u_base_url = _uf("admin_management", v_page=v_page, v_per_page=v_per_page)
    v_base_url = _uf("admin_management", u_page=u_page, u_per_page=u_per_page)

    return render_template(
        "admin_management.html",
        users=users, vendors=vendors,
        districts=UP_DISTRICTS, is_super=True,
        u_page=u_page, u_per_page=u_per_page, u_total=u_total, u_total_pages=u_total_pages, u_base_url=u_base_url,
        v_page=v_page, v_per_page=v_per_page, v_total=v_total, v_total_pages=v_total_pages, v_base_url=v_base_url,
    )

# ---- Admin CRUD ----

@app.route("/admin-mgmt/create", methods=["POST"])
@super_admin_required
def admin_mgmt_create():
    name      = request.form.get("name", "").strip()
    username  = request.form.get("username", "").strip()
    email     = request.form.get("email", "").strip()
    role      = request.form.get("role", "admin").strip()
    raw_pw    = request.form.get("password", "").strip()

    if not name or not username:
        return jsonify({"ok": False, "error": "Name and username are required."}), 400
    if User.query.filter_by(username=username).first():
        return jsonify({"ok": False, "error": "Username already exists."}), 400

    if not raw_pw:
        raw_pw = _generate_strong_password()

    user = User(
        name=name,
        username=username,
        email=email,
        password=generate_password_hash(raw_pw),
        role=role,
        is_super=False,
        created_at=datetime.now().strftime("%d-%m-%Y")
    )
    db.session.add(user)
    db.session.commit()
    logger.info("Super admin created admin user: %s (role=%s)", username, role)

    return jsonify({
        "ok":       True,
        "name":     name,
        "username": username,
        "email":    email,
        "role":     role,
        "password": raw_pw
    })

@app.route("/admin-mgmt/edit/<int:user_id>", methods=["POST"])
@super_admin_required
def admin_mgmt_edit(user_id):
    user = User.query.get_or_404(user_id)
    if user.is_super and user.id != session["user_id"]:
        return jsonify({"ok": False, "error": "Cannot edit another Super Admin."}), 403
    user.name  = request.form.get("name", user.name).strip()
    user.email = request.form.get("email", user.email).strip()
    role       = request.form.get("role", user.role).strip()
    if not user.is_super:
        user.role = role
    db.session.commit()
    logger.info("Super admin edited user: %s", user.username)
    return jsonify({"ok": True, "message": f"User '{user.username}' updated."})

@app.route("/admin-mgmt/reset/<int:user_id>", methods=["POST"])
@super_admin_required
def admin_mgmt_reset(user_id):
    user   = User.query.get_or_404(user_id)
    raw_pw = _generate_strong_password()
    user.password = generate_password_hash(raw_pw)
    db.session.commit()
    logger.info("Super admin reset password for: %s", user.username)
    return jsonify({"ok": True, "username": user.username, "password": raw_pw})

@app.route("/admin-mgmt/toggle/<int:user_id>")
@super_admin_required
def admin_mgmt_toggle(user_id):
    user = User.query.get_or_404(user_id)
    if user.id == session["user_id"]:
        flash("You cannot disable your own account.", "warning")
        return redirect(url_for("admin_management"))
    if user.is_super:
        flash("Cannot disable the Super Admin account.", "danger")
        return redirect(url_for("admin_management"))
    if user.role == "disabled":
        user.role = "admin"
        flash(f"'{user.username}' has been enabled.", "success")
    else:
        user.role = "disabled"
        flash(f"'{user.username}' has been disabled.", "warning")
    db.session.commit()
    return redirect(url_for("admin_management"))

@app.route("/admin-mgmt/delete/<int:user_id>")
@super_admin_required
def admin_mgmt_delete(user_id):
    user = User.query.get_or_404(user_id)
    if user.id == session["user_id"]:
        flash("You cannot delete your own account.", "warning")
        return redirect(url_for("admin_management"))
    if user.is_super:
        flash("Cannot delete the Super Admin account.", "danger")
        return redirect(url_for("admin_management"))
    uname = user.username
    db.session.delete(user)
    db.session.commit()
    logger.info("Super admin deleted user: %s", uname)
    flash(f"User '{uname}' deleted.", "success")
    return redirect(url_for("admin_management"))

# ---- Vendor CRUD (Super Admin) ----

@app.route("/admin-mgmt/vendor/create", methods=["POST"])
@super_admin_required
def admin_mgmt_vendor_create():
    company  = request.form.get("company_name", "").strip()
    owner    = request.form.get("owner_name", "").strip()
    district = request.form.get("district", "").strip()
    mobile   = request.form.get("mobile", "").strip()
    email    = request.form.get("email", "").strip()
    username = request.form.get("username", "").strip()
    raw_pw   = request.form.get("password", "").strip()

    if not company or not owner or not district or not mobile:
        return jsonify({"ok": False, "error": "Company name, owner name, district and mobile are required."}), 400

    if not username:
        base = owner.lower().replace(" ", ".")[:10]
        username = f"{base}.{secrets.token_hex(3)}"
        while Vendor.query.filter_by(username=username).first():
            username = f"{base}.{secrets.token_hex(3)}"
    elif Vendor.query.filter_by(username=username).first():
        return jsonify({"ok": False, "error": "Username already exists."}), 400

    if not raw_pw:
        raw_pw = _generate_strong_password()

    vendor = Vendor(
        company_name=company,
        owner_name=owner,
        mobile=mobile,
        email=email,
        username=username,
        password=generate_password_hash(raw_pw),
        district=district,
        is_active=True,
        created_at=datetime.now().strftime("%d-%m-%Y")
    )
    db.session.add(vendor)
    db.session.commit()
    logger.info("Super admin created vendor: %s", username)

    return jsonify({
        "ok":           True,
        "company_name": company,
        "owner_name":   owner,
        "username":     username,
        "password":     raw_pw,
        "district":     district,
        "email":        email,
        "mobile":       mobile
    })

@app.route("/admin-mgmt/vendor/edit/<int:vendor_id>", methods=["POST"])
@super_admin_required
def admin_mgmt_vendor_edit(vendor_id):
    vendor = Vendor.query.get_or_404(vendor_id)
    vendor.company_name = request.form.get("company_name", vendor.company_name).strip()
    vendor.owner_name   = request.form.get("owner_name", vendor.owner_name).strip()
    vendor.district     = request.form.get("district", vendor.district).strip()
    vendor.mobile       = request.form.get("mobile", vendor.mobile).strip()
    vendor.email        = request.form.get("email", vendor.email).strip()
    db.session.commit()
    logger.info("Super admin edited vendor: %s", vendor.username)
    return jsonify({"ok": True, "message": f"Vendor '{vendor.company_name}' updated."})

@app.route("/admin-mgmt/vendor/reset/<int:vendor_id>", methods=["POST"])
@super_admin_required
def admin_mgmt_vendor_reset(vendor_id):
    vendor = Vendor.query.get_or_404(vendor_id)
    raw_pw = _generate_strong_password()
    vendor.password = generate_password_hash(raw_pw)
    db.session.commit()
    logger.info("Super admin reset vendor password: %s", vendor.username)
    return jsonify({"ok": True, "username": vendor.username, "password": raw_pw})

@app.route("/admin-mgmt/vendor/toggle/<int:vendor_id>")
@super_admin_required
def admin_mgmt_vendor_toggle(vendor_id):
    vendor = Vendor.query.get_or_404(vendor_id)
    vendor.is_active = not vendor.is_active
    status = "enabled" if vendor.is_active else "disabled"
    db.session.commit()
    flash(f"Vendor '{vendor.company_name}' has been {status}.", "success")
    return redirect(url_for("admin_management") + "#vendors")

@app.route("/admin-mgmt/vendor/delete/<int:vendor_id>")
@super_admin_required
def admin_mgmt_vendor_delete(vendor_id):
    vendor = Vendor.query.get_or_404(vendor_id)
    # Unlink leads
    Lead.query.filter_by(vendor_id=vendor.id).update({"vendor_id": None})
    name = vendor.company_name
    db.session.delete(vendor)
    db.session.commit()
    logger.info("Super admin deleted vendor: %s", name)
    flash(f"Vendor '{name}' deleted.", "success")
    return redirect(url_for("admin_management") + "#vendors")

# ==========================
# LOGOUT (Admin)
# ==========================
@app.route("/logout")
def logout():
    username = session.get("username", "unknown")
    session.clear()
    logger.info("Admin logout: %s", username)
    return redirect(url_for("admin_login"))


# ==========================
# NOTIFICATIONS API
# ==========================
def push_notification(target_type, title, message, target_id=None):
    n = Notification(
        target_type=target_type, target_id=target_id,
        title=title, message=message,
        created_at=datetime.now().strftime("%d-%m-%Y %H:%M")
    )
    db.session.add(n)
    db.session.commit()


@app.route("/notifications/unread-count")
def notifications_unread_count():
    if "user_id" in session:
        count = Notification.query.filter_by(target_type="admin", is_read=False).count()
        return jsonify({"count": count})
    if "vendor_id" in session:
        count = Notification.query.filter_by(
            target_type="vendor", target_id=session["vendor_id"], is_read=False
        ).count()
        return jsonify({"count": count})
    return jsonify({"count": 0})


@app.route("/notifications/list")
def notifications_list():
    if "user_id" in session:
        items = Notification.query.filter_by(target_type="admin")\
            .order_by(Notification.id.desc()).limit(20).all()
    elif "vendor_id" in session:
        items = Notification.query.filter(
            Notification.target_type == "vendor",
            Notification.target_id == session["vendor_id"]
        ).order_by(Notification.id.desc()).limit(20).all()
    else:
        return jsonify([])
    return jsonify([{
        "id": n.id, "title": n.title, "message": n.message,
        "is_read": n.is_read, "created_at": n.created_at
    } for n in items])


@app.route("/notifications/mark-read", methods=["POST"])
def notifications_mark_read():
    data = request.get_json(silent=True) or {}
    nid  = data.get("id")
    if nid:
        n = Notification.query.get(nid)
        if n:
            n.is_read = True
            db.session.commit()
    else:
        if "user_id" in session:
            Notification.query.filter_by(target_type="admin").update({"is_read": True})
        elif "vendor_id" in session:
            Notification.query.filter_by(
                target_type="vendor", target_id=session["vendor_id"]
            ).update({"is_read": True})
        db.session.commit()
    return jsonify({"ok": True})


# ==========================
# VENDOR RATING
# ==========================
@app.route("/vendor-rate/<int:vendor_id>", methods=["POST"])
@login_required
def vendor_rate(vendor_id):
    try:
        rating = min(5.0, max(0.0, float(request.form.get("rating", 0))))
    except (ValueError, TypeError):
        rating = 0.0
    review  = request.form.get("review", "").strip()
    lead_id = request.form.get("lead_id")
    r = VendorRating(
        vendor_id=vendor_id,
        lead_id=int(lead_id) if lead_id else None,
        rating=rating, review=review,
        rated_by=session.get("username", "Admin"),
        created_at=datetime.now().strftime("%d-%m-%Y %H:%M")
    )
    db.session.add(r)
    db.session.commit()
    return jsonify({"ok": True, "message": "Rating submitted."})


# ==========================
# REVERSE AUCTION / LEAD AUCTION
# ==========================
@app.route("/lead-auction/<int:lead_id>")
@login_required
def lead_auction(lead_id):
    import json, statistics
    lead    = Lead.query.get_or_404(lead_id)
    vendors = Vendor.query.filter_by(is_active=True).all()
    eligible = []
    for v in vendors:
        cov = VendorCoverage.query.filter_by(vendor_id=v.id).first()
        district_ok = False
        if cov:
            if cov.all_districts:
                district_ok = True
            else:
                try:
                    dist_list = json.loads(cov.districts or "[]")
                except Exception:
                    dist_list = []
                district_ok = (not lead.district) or (lead.district in dist_list)
        else:
            district_ok = (not lead.district) or (v.district == lead.district)
        if not district_ok:
            continue
        ratings  = [r.rating for r in VendorRating.query.filter_by(vendor_id=v.id).all()]
        avg_rating = round(statistics.mean(ratings), 1) if ratings else 0.0
        total  = Lead.query.filter_by(vendor_id=v.id).count()
        done   = Lead.query.filter_by(vendor_id=v.id, status="Completed").count()
        conv   = round(done / total * 100, 1) if total else 0
        profile = VendorProfile.query.filter_by(vendor_id=v.id).first()
        eligible.append({
            "id": v.id, "company_name": v.company_name,
            "owner_name": v.owner_name, "district": v.district,
            "mobile": v.mobile, "email": v.email,
            "avg_rating": avg_rating, "total_leads": total,
            "completed_leads": done, "conversion_pct": conv,
            "pm_surya": profile.pm_surya_approved if profile else False,
            "discom":   profile.discom_empanelled  if profile else False,
            "years_exp": profile.years_experience  if profile else 0,
        })
    eligible.sort(key=lambda x: (-x["conversion_pct"], -x["avg_rating"]))
    return jsonify({"lead_id": lead_id, "vendors": eligible})


# ==========================
# QUOTATION COMPARISON
# ==========================
@app.route("/quotation-compare/<int:lead_id>")
@login_required
def quotation_compare(lead_id):
    import statistics
    lead   = Lead.query.get_or_404(lead_id)
    quotes = VendorQuotation.query.filter_by(lead_id=lead_id)\
                .order_by(VendorQuotation.net_price.asc()).all()
    annotated = []
    for q in quotes:
        v       = q.vendor
        ratings = [r.rating for r in VendorRating.query.filter_by(vendor_id=v.id).all()]
        avg_r   = round(statistics.mean(ratings), 1) if ratings else 0
        total   = Lead.query.filter_by(vendor_id=v.id).count()
        done    = Lead.query.filter_by(vendor_id=v.id, status="Completed").count()
        conv    = round(done / total * 100, 1) if total else 0
        profile = VendorProfile.query.filter_by(vendor_id=v.id).first()
        annotated.append({
            "q": q, "vendor": v, "avg_rating": avg_r,
            "conversion": conv, "total_leads": total,
            "profile": profile, "tag": ""
        })
    if annotated:
        min_price = min(annotated, key=lambda x: x["q"].net_price)
        max_comm  = max(annotated, key=lambda x: x["q"].commission)
        best_rate = max(annotated, key=lambda x: x["avg_rating"])
        min_price["tag"] = "💰 Best Price"
        if max_comm["q"].id != min_price["q"].id:
            max_comm["tag"] = "⭐ Highest Commission"
        if best_rate["q"].id not in (min_price["q"].id, max_comm["q"].id):
            best_rate["tag"] = "🏆 Best Rated"
    return render_template("admin_quotation_compare.html",
        lead=lead, annotated=annotated)


# ==========================
# VENDOR PERFORMANCE
# ==========================
@app.route("/vendor-performance")
@vendor_required
def vendor_performance():
    import statistics
    from collections import defaultdict
    vendor_id = session["vendor_id"]
    vendor    = Vendor.query.get_or_404(vendor_id)
    total     = Lead.query.filter_by(vendor_id=vendor_id).count()
    completed = Lead.query.filter_by(vendor_id=vendor_id, status="Completed").count()
    cancelled = Lead.query.filter_by(vendor_id=vendor_id, status="Cancelled").count()
    pending   = Lead.query.filter(
        Lead.vendor_id == vendor_id,
        Lead.status.notin_(["Completed", "Cancelled"])
    ).count()
    conv_pct  = round(completed / total * 100, 1) if total else 0
    quotes    = VendorQuotation.query.filter_by(vendor_id=vendor_id).all()
    approved_quotes   = [q for q in quotes if q.status == "Approved"]
    total_revenue     = sum(q.net_price   for q in approved_quotes)
    total_commission  = sum(q.commission  for q in approved_quotes)
    ratings  = [r.rating for r in VendorRating.query.filter_by(vendor_id=vendor_id).all()]
    avg_rating = round(statistics.mean(ratings), 1) if ratings else 0
    monthly   = defaultdict(int)
    for l in Lead.query.filter_by(vendor_id=vendor_id).all():
        if l.created_at:
            try:
                monthly[l.created_at[:7]] += 1
            except Exception:
                pass
    months_sorted = sorted(monthly.keys())[-6:]
    return render_template("vendor_performance.html",
        vendor=vendor, total=total, completed=completed,
        cancelled=cancelled, pending=pending, conv_pct=conv_pct,
        total_revenue=total_revenue, total_commission=total_commission,
        avg_rating=avg_rating, rating_count=len(ratings),
        trend_labels=months_sorted,
        trend_data=[monthly[m] for m in months_sorted],
        quotes=quotes)


# ==========================
# ADMIN REPORTS
# ==========================
@app.route("/admin/reports")
@login_required
def admin_reports():
    import statistics
    from collections import defaultdict
    vendors = Vendor.query.all()
    vendor_stats = []
    for v in vendors:
        total    = Lead.query.filter_by(vendor_id=v.id).count()
        done     = Lead.query.filter_by(vendor_id=v.id, status="Completed").count()
        cancelled= Lead.query.filter_by(vendor_id=v.id, status="Cancelled").count()
        conv     = round(done / total * 100, 1) if total else 0
        ratings  = [r.rating for r in VendorRating.query.filter_by(vendor_id=v.id).all()]
        avg_r    = round(statistics.mean(ratings), 1) if ratings else 0
        qs       = VendorQuotation.query.filter_by(vendor_id=v.id, status="Approved").all()
        revenue  = sum(q.net_price   for q in qs)
        commission=sum(q.commission  for q in qs)
        vendor_stats.append({
            "vendor": v, "total": total, "done": done,
            "cancelled": cancelled, "conv": conv,
            "avg_rating": avg_r, "revenue": revenue, "commission": commission
        })
    top_by_done  = sorted(vendor_stats, key=lambda x: -x["done"])[:10]
    top_by_rev   = sorted(vendor_stats, key=lambda x: -x["revenue"])[:10]
    top_by_comm  = sorted(vendor_stats, key=lambda x: -x["commission"])[:10]
    top_by_conv  = sorted(vendor_stats, key=lambda x: -x["conv"])[:10]
    dist_counts  = defaultdict(int)
    for l in Lead.query.all():
        if l.district:
            dist_counts[l.district] += 1
    top_districts   = sorted(dist_counts.items(), key=lambda x: -x[1])[:10]
    # Convert tuples → dicts for clean Jinja2 access
    top_districts_list = [{"name": d, "count": c} for d, c in top_districts]
    # Pre-serialize chart data as JSON for safe template injection
    import json as _json
    dist_chart_labels = _json.dumps([d["name"]  for d in top_districts_list])
    dist_chart_data   = _json.dumps([d["count"] for d in top_districts_list])
    conv_chart_labels = _json.dumps([r["vendor"].company_name for r in top_by_conv[:5]])
    conv_chart_data   = _json.dumps([r["conv"] for r in top_by_conv[:5]])
    total_leads     = Lead.query.count()
    total_vendors   = Vendor.query.count()
    active_vendors  = Vendor.query.filter_by(is_active=True).count()
    total_quotations= VendorQuotation.query.count()
    approved_count  = VendorQuotation.query.filter_by(status="Approved").count()
    all_approved    = VendorQuotation.query.filter_by(status="Approved").all()
    total_revenue   = sum(q.net_price   for q in all_approved)
    total_commission= sum(q.commission  for q in all_approved)
    return render_template("admin_reports.html",
        top_by_done=top_by_done, top_by_rev=top_by_rev,
        top_by_comm=top_by_comm, top_by_conv=top_by_conv,
        top_districts=top_districts_list,
        dist_chart_labels=dist_chart_labels, dist_chart_data=dist_chart_data,
        conv_chart_labels=conv_chart_labels, conv_chart_data=conv_chart_data,
        total_leads=total_leads, total_vendors=total_vendors,
        active_vendors=active_vendors, total_quotations=total_quotations,
        approved_count=approved_count, total_revenue=total_revenue,
        total_commission=total_commission)


# ==========================
# GLOBAL SEARCH
# ==========================
@app.route("/global-search")
@login_required
def global_search():
    q = request.args.get("q", "").strip()
    if not q:
        return jsonify({"vendors": [], "leads": [], "quotations": []})
    like = f"%{q}%"
    vendors = Vendor.query.filter(db.or_(
        Vendor.company_name.like(like), Vendor.owner_name.like(like),
        Vendor.mobile.like(like), Vendor.district.like(like),
        Vendor.username.like(like)
    )).limit(10).all()
    leads = Lead.query.filter(db.or_(
        Lead.name.like(like), Lead.phone.like(like),
        Lead.city.like(like), Lead.district.like(like)
    )).limit(10).all()
    quotes = VendorQuotation.query.filter(db.or_(
        VendorQuotation.customer_name.like(like),
        VendorQuotation.customer_phone.like(like)
    )).limit(10).all()
    return jsonify({
        "vendors":    [{"id": v.id, "name": v.company_name, "district": v.district} for v in vendors],
        "leads":      [{"id": l.id, "name": l.name, "phone": l.phone, "district": l.district} for l in leads],
        "quotations": [{"id": q.id, "customer": q.customer_name, "status": q.status} for q in quotes]
    })


# ==========================
# BLOG ROUTES
# ==========================
@app.route("/blog")
def blog_list():
    page = request.args.get('page', 1, type=int)
    blogs_query = Blog.query.filter_by(is_published=True).order_by(Blog.id.desc())
    blogs = blogs_query.paginate(page=page, per_page=12, error_out=False)
    
    # Sidebar data
    categories = db.session.query(Blog.category, db.func.count(Blog.id)).filter_by(is_published=True).group_by(Blog.category).all()
    popular = Blog.query.filter_by(is_published=True).order_by(db.func.random()).limit(5).all()
    
    return render_template("blog.html", blogs=blogs, categories=categories, popular=popular)

@app.route("/blog/<slug>")
def blog_detail(slug):
    blog = Blog.query.filter_by(slug=slug, is_published=True).first_or_404()
    related = Blog.query.filter(Blog.id != blog.id, Blog.category == blog.category, Blog.is_published == True).limit(3).all()
    return render_template("blog_detail.html", blog=blog, related=related)

@app.route("/blog/category/<category>")
def blog_category(category):
    page = request.args.get('page', 1, type=int)
    blogs_query = Blog.query.filter_by(category=category, is_published=True).order_by(Blog.id.desc())
    blogs = blogs_query.paginate(page=page, per_page=12, error_out=False)
    categories = db.session.query(Blog.category, db.func.count(Blog.id)).filter_by(is_published=True).group_by(Blog.category).all()
    popular = Blog.query.filter_by(is_published=True).order_by(db.func.random()).limit(5).all()
    return render_template("blog.html", blogs=blogs, categories=categories, popular=popular, current_category=category)

@app.route("/blog/tag/<tag>")
def blog_tag(tag):
    page = request.args.get('page', 1, type=int)
    blogs_query = Blog.query.filter(Blog.tags.like(f"%{tag}%"), Blog.is_published==True).order_by(Blog.id.desc())
    blogs = blogs_query.paginate(page=page, per_page=12, error_out=False)
    categories = db.session.query(Blog.category, db.func.count(Blog.id)).filter_by(is_published=True).group_by(Blog.category).all()
    popular = Blog.query.filter_by(is_published=True).order_by(db.func.random()).limit(5).all()
    return render_template("blog.html", blogs=blogs, categories=categories, popular=popular, current_tag=tag)

@app.route("/search")
@app.route("/blog/search")
def blog_search():
    q = request.args.get("q", "").strip()
    page = request.args.get('page', 1, type=int)
    if not q:
        return redirect(url_for('blog_list'))
    
    like_q = f"%{q}%"
    blogs_query = Blog.query.filter(db.or_(Blog.title.like(like_q), Blog.content.like(like_q), Blog.tags.like(like_q)), Blog.is_published==True).order_by(Blog.id.desc())
    blogs = blogs_query.paginate(page=page, per_page=12, error_out=False)
    categories = db.session.query(Blog.category, db.func.count(Blog.id)).filter_by(is_published=True).group_by(Blog.category).all()
    popular = Blog.query.filter_by(is_published=True).order_by(db.func.random()).limit(5).all()
    return render_template("blog.html", blogs=blogs, categories=categories, popular=popular, search_query=q)


# ==========================
# TRUST PAGES & SEO ROUTES
# ==========================
@app.route("/privacy-policy")
def privacy_policy():
    return render_template("trust/privacy.html")

@app.route("/cookie-policy")
def cookie_policy():
    return render_template("trust/cookie.html")

@app.route("/disclaimer")
def disclaimer():
    return render_template("trust/disclaimer.html")

@app.route("/terms")
def terms():
    return render_template("trust/terms.html")

@app.route("/editorial-policy")
def editorial_policy():
    return render_template("trust/editorial.html")

@app.route("/our-mission")
def our_mission():
    return render_template("trust/mission.html")

@app.route("/why-trust-us")
def why_trust_us():
    return render_template("trust/trust_us.html")

@app.route("/how-we-verify")
def how_we_verify():
    return render_template("trust/verify.html")

from flask import send_from_directory

@app.route('/googledddba798461442fe.html')
def google_verify():
    return send_from_directory('.', 'googledddba798461442fe.html')



@app.route("/sitemap.xml")
def sitemap():
    import datetime
    blogs = Blog.query.filter_by(is_published=True).all()
    today = datetime.datetime.now().strftime("%Y-%m-%d")
    
    xml = '<?xml version="1.0" encoding="UTF-8"?>\n'
    xml += '<urlset xmlns="http://www.sitemaps.org/schemas/sitemap/0.9">\n'
    
    # Static pages
    pages = ['/', '/about', '/services', '/contact', '/blog', '/privacy-policy', '/terms', '/disclaimer', '/cookie-policy', '/editorial-policy', '/our-mission', '/why-trust-us', '/how-we-verify']
    for p in pages:
        xml += '  <url>\n'
        xml += f'    <loc>https://hargharsolar.duckdns.org{p}</loc>\n'
        xml += f'    <lastmod>{today}</lastmod>\n'
        xml += '    <changefreq>weekly</changefreq>\n'
        xml += '    <priority>0.8</priority>\n'
        xml += '  </url>\n'
        
    # Blog posts
    for b in blogs:
        xml += '  <url>\n'
        xml += f'    <loc>https://hargharsolar.duckdns.org/blog/{b.slug}</loc>\n'
        date_str = b.updated_at[:10] if b.updated_at else today
        xml += f'    <lastmod>{date_str}</lastmod>\n'
        xml += '    <changefreq>monthly</changefreq>\n'
        xml += '    <priority>0.6</priority>\n'
        xml += '  </url>\n'
        
    xml += '</urlset>'
    
    from flask import Response
    return Response(xml, mimetype='application/xml')

@app.route("/rss.xml")
def rss_feed():
    import datetime
    from email.utils import formatdate
    blogs = Blog.query.filter_by(is_published=True).order_by(Blog.id.desc()).limit(20).all()
    
    xml = '<?xml version="1.0" encoding="UTF-8" ?>\n'
    xml += '<rss version="2.0" xmlns:atom="http://www.w3.org/2005/Atom">\n'
    xml += '<channel>\n'
    xml += '  <title>Har Ghar Solar Blog</title>\n'
    xml += '  <link>https://hargharsolar.duckdns.org/blog</link>\n'
    xml += '  <description>Latest solar energy news, subsidies, and guides</description>\n'
    xml += '  <language>en-in</language>\n'
    xml += '  <atom:link href="https://hargharsolar.duckdns.org/rss.xml" rel="self" type="application/rss+xml" />\n'
    
    for b in blogs:
        xml += '  <item>\n'
        xml += f'    <title>{b.title}</title>\n'
        xml += f'    <link>https://hargharsolar.duckdns.org/blog/{b.slug}</link>\n'
        xml += f'    <description><![CDATA[{b.meta_desc}]]></description>\n'
        xml += f'    <category>{b.category}</category>\n'
        # Approximate pub date
        pub_date = formatdate(timeval=None, localtime=False, usegmt=True)
        xml += f'    <pubDate>{pub_date}</pubDate>\n'
        xml += f'    <guid>https://hargharsolar.duckdns.org/blog/{b.slug}</guid>\n'
        xml += '  </item>\n'
        
    xml += '</channel>\n</rss>'
    
    from flask import Response
    return Response(xml, mimetype='application/xml')



@app.route("/solar-tools")
def solar_tools():
    return render_template("solar_tools.html")

@app.route("/downloads")
def downloads():
    return render_template("downloads.html")

# ==========================
# ERROR HANDLERS
# ==========================
@app.errorhandler(400)
def bad_request(e):
    return render_template("errors/400.html"), 400

@app.errorhandler(403)
def forbidden(e):
    return render_template("errors/403.html"), 403

@app.errorhandler(404)
def not_found(e):
    return render_template("errors/404.html"), 404

@app.errorhandler(429)
def too_many_requests(e):
    return render_template("errors/429.html"), 429

@app.errorhandler(500)
def server_error(e):
    logger.error("500 error: %s", str(e))
    return render_template("errors/500.html"), 500

@app.route("/robots.txt")
def robots():
    return send_from_directory("static", "robots.txt", mimetype="text/plain")

# ==========================
# RUN SERVER
# ==========================
if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5002))
    app.run(host="0.0.0.0", port=port, debug=False)
